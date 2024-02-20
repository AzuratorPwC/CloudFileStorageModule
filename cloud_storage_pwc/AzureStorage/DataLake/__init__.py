
from ast import main
from ..StorageAccountVirtualClass import *

from azure.storage.filedatalake import DataLakeServiceClient

from azure.identity import DefaultAzureCredential
from azure.identity import ClientSecretCredential
import pyarrow as pa
from io import BytesIO
import pyarrow.parquet as pq
import uuid
import os
import numpy as np 
import pandas as pd
from openpyxl import Workbook
from itertools import product
import polars as pl
import time
from datetime import datetime
import csv
from ..Utils import *
from ..Exceptions import *
import logging
from azure.core.exceptions import HttpResponseError,ResourceNotFoundError,ResourceExistsError

class DataLake(StorageAccountVirtualClass):
 
    def __init__(self, url:str, access_key:str=None, tenant_id:str=None, application_id:str=None,application_secret:str=None):
        super().__init__()
        try:
            if access_key is not None and tenant_id is None and application_id is None and application_secret is None:
                logging.info("DataLake-create by accesskey %s",url)
                self.__service_client = DataLakeServiceClient(account_url=url, credential=access_key)
            elif access_key is  None and tenant_id is None and application_id is None and application_secret is None: 
                logging.info("DataLake-create by defaultazurecredential %s", url)
                credential = DefaultAzureCredential()
                self.__service_client = DataLakeServiceClient(account_url=url, credential=credential)
            elif access_key is  None and tenant_id is not None and application_id is not None and application_secret is not None:
                logging.info("DataLake-create by clientsecretcredential %s", url)
                token_credential = ClientSecretCredential(
                    tenant_id, application_id,application_secret)
                self.__service_client = DataLakeServiceClient(account_url=url, credential=token_credential)
            self.__service_client.get_service_properties()
        except ResourceNotFoundError as e:
            logging.error(f"Storage account {url} not found")
            raise StorageAccountNotFound(f"Storage account {url} not found") from e
        except HttpResponseError as e:
            logging.error(f"Storage account {url} authorization error")
            raise StorageAccountAuthenticationError(f"Storage account {url} authorization error") from e
        
    def check_is_dfs(self)->bool:
        return True

    def save_binary_file(self,input_bytes:bytes, container_name:str, directory_path:str,file_name:str,is_overwrite:bool=True):
        file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
        
        main_directory_client = file_system_client.get_directory_client("/")
        if directory_path =="":
            file_client = main_directory_client.get_file_client(file_name)
        else:
            subdir_client = main_directory_client.get_sub_directory_client(directory_path)
            if subdir_client.exists() is False:
                subdir_client.create_directory()
            file_client = subdir_client.get_file_client(file_name)
        #content_settings = ContentSettings(content_encoding=encoding,content_type = "text/csv")
        file_client.upload_data(bytes(input_bytes),overwrite=is_overwrite)
    
    def read_binary_file(self, container_name:str, directory_path:str, file_name:str)->bytes:
        file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
        
        valid_file=False
        main_directory_client = file_system_client.get_directory_client("/")
        if directory_path =="":
            file_client = main_directory_client.get_file_client(file_name)
            if  file_client.exists():
                valid_file=True
        else:
            subdir_client = main_directory_client.get_sub_directory_client(directory_path)
            if subdir_client.exists() is True:
                file_client = subdir_client.get_file_client(file_name)
                if file_client.exists():
                    valid_file=True
            
        if valid_file:
            download = file_client.download_file()
            download_bytes = download.readall()
        else:
            raise BlobNotFound(f"{container_name}/{directory_path}/{file_name} not found")
        return download_bytes

    

        
    
    def read_excel_file(self,container_name:str,directory_path:str,file_name:str,engine: ENGINE_TYPES ='polars',skip_rows:int = 0,is_first_row_as_header:bool = False,sheets:list=None,tech_columns:bool=False):
        file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
        directory_client = file_system_client.get_directory_client(directory_path)
        file_client = directory_client.get_file_client(file_name)
        download = file_client.download_file()
        download_bytes = download.readall()
        if file_name.endswith(".xls"):
            workbook = pd.ExcelFile(download_bytes, engine='xlrd') 
        else:
            workbook = pd.ExcelFile(download_bytes)
          
        workbook_sheetnames = workbook.sheet_names # get all sheet names
        if bool(sheets):
            workbook_sheetnames=list(set(workbook_sheetnames) & set(sheets))
        list_of_dff = []
        if is_first_row_as_header:
            is_first_row_as_header=0
        else:
            is_first_row_as_header=None
        for sheet in workbook_sheetnames:
            dff = pd.read_excel(workbook, sheet_name = sheet,skip_rows=skip_rows, index_col = None, header = is_first_row_as_header)
            if tech_columns:
                df =  add_tech_columns(df,container_name,directory_path.replace("\\","/"),file_name)
            
            list_of_dff.append(DataFromExcel(dff,sheet))
            
        return list_of_dff
    
    
    def delete_file(self,container_name : str,directory_path : str,file_name:str,wait:bool=True):
        try:
            file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
            if file_system_client.exists() is False:
                raise ContainerNotFound(f"Container {container_name} not found")
            
            main_directory = file_system_client.get_directory_client("/")
            if directory_path != "":
                main_directory = main_directory.get_sub_directory_client(directory_path)
                #if main_directory.exists() is False:
                #    pass
            
            file_client= main_directory.get_file_client(file_name)
            if file_client.exists():
                file_client.delete_file()
                   
            if wait:
                check_if_exist = file_client.exists()
                while check_if_exist:
                    time.sleep(1)
                    check_if_exist = file_client.exists()
        except ContainerNotFound as e:
            raise e
        except Exception as e:
            raise Exception(f"Error listing files in container {container_name}") from e
                            
                            
                            
    def delete_files_by_prefix(self,container_name : str,directory_path : str,file_prefix:str, recursive:bool=False,wait:bool=True):
        try:
            file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
            if file_system_client.exists() is False:
                raise ContainerNotFound(f"Container {container_name} not found")
            
            files_exists = self.ls_files(container_name,directory_path,recursive)
            files = [f for f in files_exists if f.split("/")[-1].startswith(file_prefix)]
            for f in files:
                self.delete_file(container_name,f.removesuffix(f"/{f.split('/')[-1]}"),f.split("/")[-1])

            if wait and files:
                time.sleep(1)
                files_exists = self.ls_files(container_name,directory_path,recursive)
                files = [f for f in files_exists if f.split("/")[-1].startswith(file_prefix)]
                while files:
                    time.sleep(1)
                    files_exists = self.ls_files(container_name,directory_path,recursive)
                    files = [f for f in files_exists if f.split("/")[-1].startswith(file_prefix)]
        except ContainerNotFound as e:
            raise e
        except Exception as e:
            raise Exception(f"Error listing files in container {container_name}") from e
        
    def delete_folder(self,container_name : str,directory_path : str,wait:bool=True):
        try:
            file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
            if file_system_client.exists() is False:
                raise ContainerNotFound(f"Container {container_name} not found")
            
            main_directory = file_system_client.get_directory_client("/")
            if directory_path != "":
                main_directory = main_directory.get_sub_directory_client(directory_path)

            main_directory.delete_directory()
            
            if wait :
                time.sleep(1)
                dict_exists = main_directory.exists()
                while dict_exists:
                    time.sleep(2)
                    dict_exists = main_directory.exists()
        except ContainerNotFound as e:
            raise e
        except Exception as e:
            raise Exception(f"Error listing files in container {container_name}") from e
    
    def file_exists(self, container_name : str,directory_path : str,file_name:str):
        
        file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
        if file_system_client.exists() is False:
                raise ContainerNotFound(f"Container {container_name} not found")
        
        main_directory = file_system_client.get_directory_client("/")
        if directory_path != "":
            main_directory = main_directory.get_sub_directory_client(directory_path)

        return main_directory.get_file_client(file_name).exists()
            
    def folder_exists(self, container_name : str, directory_path : str):
        
        file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
        if file_system_client.exists() is False:
                raise ContainerNotFound(f"Container {container_name} not found")
        
        main_directory = file_system_client.get_directory_client("/")
        if directory_path != "":
            main_directory = main_directory.get_sub_directory_client(directory_path)

        return main_directory.exists()
                    
    
    def move_file(self,container_name : str,directory_path : str,file_name:str,new_container_name : str,new_directory_path : str,is_overwrite :bool=True,is_delete_source_file:bool=False):
        try:
            file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
            if file_system_client.exists() is False:
                raise ContainerNotFound(f"Container {container_name} not found")
            file_system_client = self.__service_client.get_file_system_client(file_system=new_container_name)
            if file_system_client.exists() is False:
                raise ContainerNotFound(f"Container {new_container_name} not found")
            
            
            self.save_binary_file(self.read_binary_file(container_name,directory_path,file_name),new_container_name,new_directory_path,file_name,is_overwrite)
            if is_delete_source_file:
                self.delete_file(container_name ,directory_path ,file_name)
            
        except ContainerNotFound as e:
            raise e
        except Exception as e:
            raise Exception(f"Error listing files in container {container_name}") from e
    
    def ls_files(self, container_name:str, directory_path:str, recursive:bool=False)->list:
        try:
            file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
            if file_system_client.exists() is False:
                raise ContainerNotFound(f"Container {container_name} not found")

            if directory_path=="":
                directory_path="/"
            files = []
            generator = file_system_client.get_paths(path=directory_path, recursive=recursive)
            for file in generator:
                if file.is_directory is False:
                    files.append(file.name)
            return files
        except ContainerNotFound as e:
            raise e
        except Exception as e:
            raise Exception(f"Error listing files in container {container_name}") from e
        
        
        
    #def save_dataframe_as_xlsx(self, df,container_name : str,directory_path:str ,file_name:str,sheet_name:str,engine:ENGINE_TYPES ='polars',index=False,header=False):
    #    super().save_dataframe_as_xlsx(df,container_name,directory_path,file_name,sheet_name,engine,index,header)
        
    
    def move_folder(self,container_name : str,directory_path : str,new_container_name : str,new_directory_path : str,is_overwrite :bool=True,is_delete_source_folder:bool=False)->bool:
        try:
            file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
            if file_system_client.exists() is False:
                raise ContainerNotFound(f"Container {container_name} not found")
            file_system_client = self.__service_client.get_file_system_client(file_system=new_container_name)
            if file_system_client.exists() is False:
                raise ContainerNotFound(f"Container {new_container_name} not found")

            
            files_exists = self.ls_files(container_name,directory_path,True)
            
            for f in files_exists:
                self.move_file(container_name,directory_path,f"{f.removeprefix(directory_path)}/",new_container_name,new_directory_path,is_overwrite,is_delete_source_folder)

        except ContainerNotFound as e:
            raise e
        except Exception as e:
            raise Exception(f"Error listing files in container {container_name}") from e
        
    def rename_file(self,container_name : str,directory_path : str,file_name:str,new_file_name:str):
        try:
            file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
            if file_system_client.exists() is False:
                raise ContainerNotFound(f"Container {container_name} not found")
            main_directory = file_system_client.get_directory_client("/")
            if directory_path != "":
                main_directory = main_directory.get_sub_directory_client(directory_path)
                
            main_directory.get_file_client(file_name).rename_file(f"{container_name}/{directory_path}/{new_file_name}") 

        except ContainerNotFound as e:
            raise e
        except Exception as e:
            raise Exception(f"Error listing files in container {container_name}") from e
    
    def rename_folder(self,container_name : str,directory_path : str,new_directory_path:str):
        try:
            file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
            if file_system_client.exists() is False:
                raise ContainerNotFound(f"Container {container_name} not found")
            
            main_directory = file_system_client.get_directory_client("/")
            if directory_path != "":
                main_directory = main_directory.get_sub_directory_client(directory_path)
                
            main_directory.rename_directory(f"{container_name}/{new_directory_path}")

        except ContainerNotFound as e:
            raise e
        except Exception as e:
            raise Exception(f"Error listing files in container {container_name}") from e
        
    def create_empty_file(self,container_name : str,directory_path : str,file_name:str):
        self.save_binary_file(b' ',container_name,directory_path,file_name)
        
    def create_container(self,container_name : str,public_access:CONTAINER_ACCESS_TYPES ='Private'):
        try:
            container_client = self.__service_client.get_file_system_client(file_system=container_name).get_file_client("aaa")
            if not container_client.exists():
                if public_access == "Private":
                    public_access_ = None
                else:
                    public_access_ = None
                self.__service_client.create_file_system(file_system=container_name,public_access=public_access_)
        except ResourceExistsError as e:
            raise ContainerAccessTypes(f"Container access {public_access} is not allowe in {container_name}") from e
        except Exception as e:
            raise Exception(f"Error creating container {container_name}") from e

        
    def delete_container(self,container_name : str):
        try:
            container_client = self.__service_client.get_file_system_client(file_system=container_name)
            if container_client.exists():
                container_client.delete_file_system()
        except Exception as e:
            raise Exception(f"Error deleting container {container_name}") from e