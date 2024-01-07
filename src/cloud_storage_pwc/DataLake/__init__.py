
from ..StorageAccountVirtualClass import *
from ..StorageAccountVirtualClass import __addTechColumns__
from azure.storage.filedatalake import DataLakeServiceClient, ContentSettings

from azure.identity import DefaultAzureCredential
from azure.identity import ClientSecretCredential
import pyarrow as pa
from io import BytesIO
import pyarrow.parquet as pq
import uuid
import os
import numpy as np 
import pandas as pd
from openpyxl import Workbook
from itertools import product
import polars as pl
import time
from datetime import datetime
import csv


class DataLake(StorageAccountVirtualClass):
 
    def __init__(self, url:str,accessKey:str=None,tenantId:str=None,applicationId:str=None,applicationSecret:str=None):
        super().__init__()
        try:
            if accessKey is not None and tenantId is None and applicationId is None and applicationSecret is None: 
                self.__service_client = DataLakeServiceClient(account_url=url, credential=accessKey)
            elif accessKey is  None and tenantId is None and applicationId is None and applicationSecret is None: 
                credential = DefaultAzureCredential()
                self.__service_client = DataLakeServiceClient(account_url=url, credential=credential)
            elif accessKey is  None and tenantId is not None and applicationId is not None and applicationSecret is not None:
                token_credential = ClientSecretCredential(
                tenantId, #self.active_directory_tenant_id,
                applicationId, #self.active_directory_application_id,
                applicationSecret #self.active_directory_application_secret
                )
                self.__service_client = DataLakeServiceClient(account_url=url, credential=token_credential)
        except Exception as e:
            if "getaddrinfo failed" in str(e):
                raise Exception(f"Warning: Storage account not found.")
            else:
                raise Exception(f"Blad logowania na {url}: {str(e)}")
        
        
        #print(self.__service_client.)
        
        #if bool(self.__service_client.get_account_information()['is_hns_enabled']) =='False':
        #    raise Exception(f"storage {url} nie jest datalake")
        
    def _check_is_blob(self):

       
        return False

    

    def save_binary_file(self, inputbytes:bytes,containerName : str,directoryPath : str,fileName:str,sourceEncoding: StorageAccountVirtualClass._ENCODING_TYPES = "UTF-8",isOverWrite :bool=True):
        file_system_client = self.__service_client.get_file_system_client(file_system=containerName)        
        directory_client = file_system_client.get_directory_client(directoryPath)
        new_file_client = directory_client.create_file(fileName)
        #content_settings = ContentSettings(content_encoding=sourceEncoding,content_type = "text/csv")
        new_file_client.upload_data(bytes(inputbytes),overwrite=isOverWrite)
    
    def read_binary_file(self,containerName : str,directoryPath : str,fileName:str)->bytes:
        file_system_client = self.__service_client.get_file_system_client(file_system=containerName) 
        directory_client = file_system_client.get_directory_client(directoryPath)
        file_client = directory_client.get_file_client(fileName)
        download = file_client.download_file()
        download_bytes = download.readall()
        return download_bytes

    
    def read_csv_file(self,containerName:str,directoryPath:str,sourceFileName:str,engine: StorageAccountVirtualClass._ENGINE_TYPES ='polars',sourceEncoding:StorageAccountVirtualClass._ENCODING_TYPES = "UTF-8", columnDelimiter:str = ";",isFirstRowAsHeader:bool = False,skipRows:int=0,skipBlankLines = True,addStrTechCol:bool=False):
        file_system_client = self.__service_client.get_file_system_client(file_system=containerName) 
        directory_client = file_system_client.get_directory_client(directoryPath)
        file_client = directory_client.get_file_client(sourceFileName)
        download = file_client.download_file()
        download_bytes = download.readall()
        df = self.read_csv_bytes(download_bytes,engine,sourceEncoding,columnDelimiter,isFirstRowAsHeader,skipRows,skipBlankLines)
        
        if addStrTechCol:
            df =  __addTechColumns__(df,containerName,directoryPath.replace("\\","/"),sourceFileName)
        return df
    
    def read_csv_folder(self,containerName:str,directoryPath:str,engine: StorageAccountVirtualClass._ENGINE_TYPES ='polars',includeSubfolders:list=None,sourceEncoding :StorageAccountVirtualClass._ENCODING_TYPES= "UTF-8", columnDelimiter:str = ";",isFirstRowAsHeader:bool = False,skipRows:int=0,skipBlankLines=True,addStrTechCol:bool=False,recursive:bool=False) ->pd.DataFrame:
        file_system_client = self.__service_client.get_file_system_client(file_system=containerName) 
        listFiles = file_system_client.get_paths(directoryPath,recursive)
        df = pd.DataFrame()
        if listFiles:
            for i,r in enumerate(listFiles):
                file = r.name.replace(directoryPath + "/","")
                dfNew = self.read_csv_file(containerName,directoryPath,file,engine,sourceEncoding,columnDelimiter,isFirstRowAsHeader,skipRows,skipBlankLines,addStrTechCol)
                df = pd.concat([df, dfNew], axis=0, join="outer", ignore_index=True)
        return df
    
    def read_excel_file(self,containerName:str,directoryPath:str,sourceFileName:str,engine: StorageAccountVirtualClass._ENGINE_TYPES ='polars',skipRows:int = 0,isFirstRowAsHeader:bool = False,sheets:list()=None,addStrTechCol:bool=False):
        file_system_client = self.__service_client.get_file_system_client(file_system=containerName) 
        directory_client = file_system_client.get_directory_client(directoryPath)
        file_client = directory_client.get_file_client(sourceFileName)
        download = file_client.download_file()
        download_bytes = download.readall()
        if sourceFileName.endswith(".xls"):
            workbook = pd.ExcelFile(download_bytes, engine='xlrd') 
        else:
            workbook = pd.ExcelFile(download_bytes)
          
        workbook_sheetnames = workbook.sheet_names # get all sheet names
        if bool(sheets):
            workbook_sheetnames=list(set(workbook_sheetnames) & set(sheets))
        list_of_dff = []
        if isFirstRowAsHeader:
            isFirstRowAsHeader=0
        else:
            isFirstRowAsHeader=None
        for sheet in workbook_sheetnames:
            dff = pd.read_excel(workbook, sheet_name = sheet,skiprows=skipRows, index_col = None, header = isFirstRowAsHeader)
            if addStrTechCol:
                df =  __addTechColumns__(df,containerName,directoryPath.replace("\\","/"),sourceFileName)
            
            list_of_dff.append(DataFromExcel(dff,sheet))
            
        return list_of_dff
    
    
    
    def save_dataframe_as_parquet(self,df:pd.DataFrame,containerName : str,directoryPath:str,partitionCols:list=None,compression:str=None):
        
        VALID_compression = {'snappy', 'gzip', 'brotli', None}
        if compression is not None:
            compression=compression.lower()
        if compression not in VALID_compression:
            raise ValueError("results: status must be one of %r." % VALID_compression)
        if not(df.empty):
            df = df.replace('\n', ' ', regex=True)
        file_system_client = self.__service_client.get_file_system_client(file_system=containerName) 
        buf = BytesIO()
        df.to_parquet(buf,allow_truncated_timestamps=True, use_deprecated_int96_timestamps=True,partition_cols=partitionCols)
        buf.seek(0)
        directory_client = file_system_client.get_directory_client(directoryPath)
        new_client_parq = directory_client.create_file(f"{uuid.uuid4().hex}.parquet")
        new_client_parq.upload_data(buf.getvalue(),overwrite=True)
    

    def save_dataframe_as_csv(self,df:[pd.DataFrame, pl.DataFrame],containerName : str,directoryPath:str,file:str=None,partitionCols:list=None,sourceEncoding:StorageAccountVirtualClass._ENCODING_TYPES= "UTF-8", columnDelimiter:str = ";",isFirstRowAsHeader:bool = True,quoteChar:str=' ',quoting:['never', 'always', 'necessary']='never',escapeChar:str="\\", engine: StorageAccountVirtualClass._ENGINE_TYPES ='polars'):
        
        quoting_dict = {'never':csv.QUOTE_NONE, 'always':csv.QUOTE_ALL, 'necessary':csv.QUOTE_MINIMAL}
        
        if isinstance(df, pd.DataFrame):
            if df.empty:
                return
            df = df.replace(r'\\n', '', regex=True) 
            if engine != 'pandas':
                df = pl.from_pandas(df)

        elif isinstance(df, pl.DataFrame):
            if df.is_empty():
                return
            df = df.with_columns(pl.col(pl.Utf8).str.replace_all(r"\\n", ""))
            if engine != 'polars':
                df = df.to_pandas(df)

                        
        if partitionCols:
            partitionDict = {}
            for x in partitionCols:
                partitionDict[x] = df[x].unique()
    
            partitionGroups = [dict(zip(partitionDict.keys(),items)) for items in product(*partitionDict.values())]
            partitionGroups = [d for d in partitionGroups if np.nan not in d.values()]

            for d in partitionGroups:
                df_part = df
                partitionPath=[]
                

                if isinstance(df_part, pd.DataFrame):
                    for d1 in d:
                        df_part = df_part[df_part[d1] == d[d1]]
                        partitionPath.append(f"{d1}={str(d[d1])}")

                if isinstance(df_part, pl.DataFrame):
                    for d1 in d:
                        df_part = df_part.filter(df_part[d1] == d[d1])
                        partitionPath.append(f"{d1}={str(d[d1])}")

                if isinstance(df_part, pd.DataFrame):
                    if not(df_part.empty):
                        buf = BytesIO()
                        df_reset = df_part.reset_index(drop=True)
                        df_reset.to_csv(buf,index=False, sep=columnDelimiter,encoding=sourceEncoding,header=isFirstRowAsHeader,quotechar=quoteChar, quoting=quoting_dict[quoting],escapechar=escapeChar)
                        buf.seek(0)
                        self.save_binary_file(buf.getvalue(),containerName ,directoryPath +"/" +"/".join(partitionPath),f"{uuid.uuid4().hex}.csv",True)


                if isinstance(df_part, pl.DataFrame):
                    if not(df_part.is_empty()):
                        buf = BytesIO()
                        df_reset = df_part
                        df_reset.write_csv(buf, separator=columnDelimiter, has_header=isFirstRowAsHeader, quote_char='"', quote_style=quoting)
                        buf.seek(0)
                        self.save_binary_file(buf.getvalue(),containerName ,directoryPath +"/" +"/".join(partitionPath),f"{uuid.uuid4().hex}.csv",True)
                                                
        
        else:
            buf = BytesIO()
            
            if isinstance(df, pd.DataFrame):
                df_reset = df.reset_index(drop=True)
                df_reset.to_csv(buf,index=False, sep=columnDelimiter,encoding=sourceEncoding,header=isFirstRowAsHeader,quotechar=quoteChar, quoting=quoting_dict[quoting],escapechar=escapeChar)
            else:
                df_reset = df
                df_reset.write_csv(buf, separator=columnDelimiter, has_header=isFirstRowAsHeader, quote_style=quoting)
    
            buf.seek(0)

            if file:
                filename = file
            else:
                filename = f"{uuid.uuid4().hex}.csv"
            self.save_binary_file(buf.getvalue(),containerName ,directoryPath,filename,True)

        
    def save_dataframe_as_parqarrow(self,df:pd.DataFrame,containerName : str,directoryPath:str,partitionCols:list=None,compression:str='NONE'):
        VALID_compression = {'NONE','SNAPPY', 'GZIP', 'BROTLI', 'LZ4', 'ZSTD'}
        if compression is not None:
            compression=compression.upper()
        elif compression is None:
            compression = 'NONE'
        if not(df.empty):
            df = df.replace('\n', ' ', regex=True)    
        if compression not in VALID_compression:
            raise ValueError("results: status must be one of %r." % VALID_compression)
        
        file_system_client = self.__service_client.get_file_system_client(file_system=containerName) 
        
        table = pa.Table.from_pandas(df)
        buf = pa.BufferOutputStream()
        pq.write_table(table, buf,use_deprecated_int96_timestamps=True, allow_truncated_timestamps=True,compression=compression)
        directory_client = file_system_client.get_directory_client(directoryPath)
        new_client_parq = directory_client.create_file(f"{uuid.uuid4().hex}.parquet")
        new_client_parq.upload_data(buf.getvalue().to_pybytes(),overwrite=True)


    def save_json_file(self, df:[pd.DataFrame, pl.DataFrame], containerName: str, directory: str, file:str = None, engine: StorageAccountVirtualClass._ENGINE_TYPES ='polars', orient: StorageAccountVirtualClass._ORIENT_TYPES = 'records'):    

        if isinstance(df, pd.DataFrame):
            if not(df.empty):
                return
            
        if isinstance(df, pl.DataFrame):
            if not(df.is_empty):
                return    
                 
        if isinstance(df, pd.DataFrame):
            buf = BytesIO()           
            buf.seek(0)
            if engine == 'pandas':           
                df.to_json(buf, orient=orient)
            else:
                polars_df = pl.from_pandas(df)
                polars_df.write_json(buf, row_oriented=(orient=='records'), pretty=True)

        elif isinstance(df, pl.DataFrame):
            buf = BytesIO()           
            buf.seek(0)
            if engine == 'pandas':
                pandas_df = df.to_pandas()
                pandas_df.to_json(buf, orient=orient)
            else:
                df.write_json(buf, row_oriented=(orient=='records'), pretty=True)
        else:
            raise Exception("DF argument is not Pandas or Polars DataFrame")

        if file:
            filename = file
        else:
            filename = f"{uuid.uuid4().hex}.json"
        self.save_binary_file(buf.getvalue(), containerName ,directory,filename,True)

    def save_listdataframe_as_xlsx(self,list_df:list, list_sheetnames:list, containerName : str,directoryPath:str ,fileName:str,index=False,header=False):
        # for multiple dfs
        buf = BytesIO() 
        with pd.ExcelWriter(buf, engine = 'openpyxl') as writer:
            for df, sheet_name in zip(list_df,list_sheetnames):
                df.to_excel(writer, index = index, header = header, sheet_name = sheet_name)   
        
        file_system_client = self.__service_client.get_file_system_client(file_system=containerName) 
        directory_client = file_system_client.get_directory_client(directoryPath)
        new_client_xlsx = directory_client.create_file(f"{fileName}.xlsx")
        buf.seek(0)
        new_client_xlsx.upload_data(buf.getvalue(),overwrite=True)
    
    
    def read_parquet_file(self, containerName: str, directoryPath: str,sourceFileName:str, columns: list = None,addStrTechCol:bool=False)->pd.DataFrame:
        file_system_client = self.__service_client.get_file_system_client(file_system=containerName)
        directory_client = file_system_client.get_directory_client(directoryPath)
        file_client= directory_client.get_file_client(sourceFileName)
        download = file_client.download_file()
        download_bytes = download.readall()
        df = self.read_parquet_bytes(bytes=download_bytes,columns=columns)
        if addStrTechCol:
            df =  __addTechColumns__(df,containerName,directoryPath.replace("\\","/"),sourceFileName)
        
        return df
        
        
    
    def read_parquet_folder(self,containerName:str,directoryPath:str,includeSubfolders:list=None,columns:list = None,addStrTechCol:bool=False,recursive:bool=False) ->pd.DataFrame:
        file_system_client = self.__service_client.get_file_system_client(file_system=containerName)
        #directory_client = file_system_client.get_directory_client(directoryPath)
        listFiles = file_system_client.get_paths(directoryPath,recursive)
        df = pd.DataFrame()
        if listFiles:
            for i,r in enumerate(listFiles):
                file = str(r.name).replace(directoryPath + "/","")
                dfNew = self.read_parquet_file(containerName,directoryPath,file,columns,addStrTechCol)
                df = pd.concat([df, dfNew], axis=0, join="outer", ignore_index=True)
        return df
    
    def delete_file(self,containerName : str,directoryPath : str,fileName:str,wait:bool=True):
        file_system_client = self.__service_client.get_file_system_client(file_system=containerName)
        
        if directoryPath== "":
            file_client= file_system_client.get_file_client(fileName)
            if file_client.exists():
                file_client.delete_file()
                
                if wait:
                    checkIfExist = file_client.exists()
                    while checkIfExist:
                        time.sleep(5)
                        checkIfExist = file_client.exists()
        else:
            directory_client = file_system_client.get_directory_client(directoryPath)
            if directory_client.exists():
                file_client= directory_client.get_file_client(fileName)
                if file_client.exists():
                    file_client.delete_file()
                    
                    if wait:
                        file_client= directory_client.get_file_client(fileName)
                        checkIfExist = file_client.exists()
                        while checkIfExist:
                            time.sleep(5)
                            file_client= directory_client.get_file_client(fileName)
                            checkIfExist = file_client.exists()
        
        
    def delete_folder(self,containerName : str,directoryPath : str,wait:bool=True):
        file_system_client = self.__service_client.get_file_system_client(file_system=containerName)
        if directoryPath== "":
            files = file_system_client.get_paths()
            for file in files:
                file.delete_file()
                
            if wait:
                files = list(file_system_client.get_paths())
                while len(files)>0:
                    files = file_system_client.get_paths()      
        else:
                 
            directory_client = file_system_client.get_directory_client(directoryPath)
            if directory_client.exists():
                directory_client.delete_directory()
        
            if wait:
                directory_client = file_system_client.get_directory_client(directoryPath)
                checkIfExist=directory_client.exists()
                while checkIfExist:
                    time.sleep(5)
                    directory_client = file_system_client.get_directory_client(directoryPath)
                    checkIfExist=directory_client.exists()
    
    def file_exists(self, containerName : str,directoryPath : str,fileName:str):
        file_system_client = self.__service_client.get_file_system_client(file_system=containerName)
        
        if directoryPath== "":
            file_client= file_system_client.get_file_client(fileName)
        else:
            directory_client = file_system_client.get_directory_client(directoryPath)
            if directory_client.exists():
                file_client= directory_client.get_file_client(fileName)
        return file_client.exists()
            
    def folder_exists(self, containerName : str, directoryPath : str):
        if directoryPath== "":
            return False       
        file_system_client = self.__service_client.get_file_system_client(file_system=containerName)
        directory_client = file_system_client.get_directory_client(directoryPath)
        return directory_client.exists()
                    
    
    def move_file(self,containerName : str,directoryPath : str,fileName:str,newContainerName : str,newDirectoryPath : str,newFileName:str,isOverWrite :bool=True,isDeleteSourceFile:bool=False):
        if not self.file_exists(containerName, directoryPath, fileName):
            raise FileNotFoundError
        self.save_binary_file(self.read_binary_file(containerName,directoryPath,fileName),newContainerName,newDirectoryPath,newFileName,isOverWrite)
        if isDeleteSourceFile:
            self.delete_file(containerName ,directoryPath ,fileName)
    
    def ls_files(self,containerName : str, directoryPath : str, recursive:bool=False)->list:
        """
        List files under a path, optionally recursively
        """
        if directoryPath=="":
            directoryPath="/"
        files = []
        file_system_client = self.__service_client.get_file_system_client(file_system=containerName)
        generator = file_system_client.get_paths(path=directoryPath, recursive=recursive)
        for file in generator:
            if file.is_directory==False:
                files.append(file.name)
        
        return files
    
    def move_folder(self,containerName : str,directoryPath : str,newContainerName : str,newDirectoryPath : str,isOverWrite :bool=True,isDeleteSourceFolder:bool=False)->bool:
        if not self.folder_exists(containerName, directoryPath):
            raise FileNotFoundError
        listFiles = self.ls_files(containerName,directoryPath,True)
        for i in listFiles:
            self.move_file(containerName,directoryPath,i,newContainerName,newDirectoryPath,i,isOverWrite,isDeleteSourceFolder)
        return True
        
    def renema_file(self,containerName : str,directoryPath : str,fileName:str,newFileName:str):
        self.move_file(containerName,directoryPath,fileName,containerName,directoryPath,newFileName,False,True)
    
    def renema_folder(self,containerName : str,directoryPath : str,newDirectoryPath:str):
        self.move_folder(containerName,directoryPath,containerName,newDirectoryPath,False,True)
        
    def create_empty_file(self,containerName : str,directoryPath : str,fileName:str):
        file_system_client = self.__service_client.get_file_system_client(file_system=containerName)
        directory_client = file_system_client.get_directory_client(directoryPath)
        file_client= directory_client.get_file_client(fileName)
        file_client.create_file()
        file_client.upload_data('')