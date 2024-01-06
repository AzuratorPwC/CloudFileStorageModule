


from ..StorageAccountVirtualClass import *
from azure.storage.blob import BlobServiceClient , ContentSettings  #, BlobClient, ContainerClient

from azure.identity import DefaultAzureCredential
from azure.identity import ClientSecretCredential
import pyarrow as pa
import pyarrow.parquet as pq
import uuid
import os
from io import BytesIO
import numpy as np 
import pandas as pd
import polars as pl
from openpyxl import Workbook
from itertools import product
import time
import csv

class Blob(StorageAccountVirtualClass):
 
    def __init__(self, url:str,accessKey:str=None,tenantId:str=None,applicationId:str=None,applicationSecret:str=None):
        super().__init__()
        try:
            if accessKey is not None and tenantId is None and applicationId is None and applicationSecret is None: 
                self.__service_client = BlobServiceClient(account_url=url, credential=accessKey)
            elif accessKey is  None and tenantId is None and applicationId is None and applicationSecret is None: 
                credential = DefaultAzureCredential()
                self.__service_client = BlobServiceClient(account_url=url, credential=credential)
            elif accessKey is  None and tenantId is not None and applicationId is not None and applicationSecret is not None:
                token_credential = ClientSecretCredential(
                tenantId, #self.active_directory_tenant_id,
                applicationId, #self.active_directory_application_id,
                applicationSecret #self.active_directory_application_secret
                )
                self.__service_client = BlobServiceClient(account_url=url, credential=token_credential)
        except Exception as e:
            if "getaddrinfo failed" in str(e):
                raise Exception(f"Warning: Storage account not found.")
            else:
                raise Exception(f"Blad logowania na {url}: {str(e)}")
        
        #if self.__service_client.get_account_information()['is_hns_enabled'] =='True':
        #    raise Exception(f"storage {url} nie jest blob")

    
    def _check_is_blob(self):

        account_info= self.__service_client.get_account_information()
        return account_info['account_kind'] == 'StorageV2' and account_info['is_hns_enabled']
            

    def ls_files(self,containerName : str, directoryPath : str, recursive:bool=False)->list:
        """
        List files under a path, optionally recursively
        """
        if not directoryPath == '' and not directoryPath.endswith('/'):
            directoryPath += '/'

        container_client = self.__service_client.get_container_client(container=containerName) 
        blob_iter = container_client.list_blobs(name_starts_with=directoryPath)
        files = []
        for blob in blob_iter:
            relative_path = os.path.relpath(blob.name, directoryPath)
            if recursive or not '\\' in relative_path:
                files.append(relative_path)
        return files
    
    def read_binary_file(self,containerName : str,directoryPath : str,fileName:str)->bytes:
        """
        dev1
        """
        if not directoryPath == '' and not directoryPath.endswith('/'):
            directoryPath += '/'
        container_client = self.__service_client.get_container_client(container=containerName)
        path = directoryPath + fileName
        blob_client = container_client.get_blob_client(path)
        download = blob_client.download_blob()
        download_bytes = download.readall()
        return download_bytes

    def save_binary_file(self, inputbytes:bytes,containerName : str,directoryPath : str,fileName:str,sourceEncoding:str = "UTF-8",isOverWrite :bool=True):
        container_client = self.__service_client.get_container_client(container=containerName)        
        
        if not directoryPath == '' and not directoryPath.endswith('/'):
            directoryPath += '/'
            
        new_blob_client = container_client.get_blob_client(directoryPath +fileName)
        #content_settings = ContentSettings(content_encoding=sourceEncoding,content_type = "text/csv")
        new_blob_client.upload_blob(bytes(inputbytes),overwrite=isOverWrite)
        
    def read_csv_file(self,containerName:str,directoryPath:str,sourceFileName:str,engine:('pandas','polars') ='pandas',sourceEncoding:str = "UTF-8", columnDelimiter:str = ";",isFirstRowAsHeader:bool = False,skipRows:int=0,skipBlankLines = True,addStrTechCol:bool=False):
        
        if not directoryPath == '' and not directoryPath.endswith('/'):
            path = directoryPath +"/"+sourceFileName
        else:
            path = directoryPath + sourceFileName
        
        
        download_bytes = self.read_binary_file(containerName,directoryPath,sourceFileName)
        if engine=='pandas':
            df = self.read_csv_bytes(download_bytes,engine,sourceEncoding,columnDelimiter,isFirstRowAsHeader,skipRows,skipBlankLines)

        elif engine=='polars':
            df = self.read_csv_bytes(download_bytes,engine,sourceEncoding,columnDelimiter,isFirstRowAsHeader,skipRows,skipBlankLines)
        if addStrTechCol:
            df['techContainer'] = containerName
            path = path.replace("\\","/")
            df['techFolderPath'] = path[0:path.rfind("/")]
            df['techSourceFile'] = path[path.rfind("/")+1:len(path)]
        return df
    
    def read_csv_folder(self,containerName:str,directoryPath:str,engine:('pandas','polars') ='pandas',includeSubfolders:list=None,sourceEncoding :str= "UTF-8", columnDelimiter:str = ";",isFirstRowAsHeader:bool = False,skipRows:int=0,skipBlankLines=True,addStrTechCol:bool=False,recursive:bool=False) ->pd.DataFrame:
        listFiles = self.ls_files(containerName,directoryPath, recursive=recursive)
        
        if  includeSubfolders:
            newList=[]
            for i in includeSubfolders:
                for j in listFiles:
                    if  j.startswith(i.replace("/","\\")):
                        newList.append(j)
                listFiles = list(set(listFiles) - set(newList))
            listFiles = newList
        
        df = pd.DataFrame()
        if listFiles:
            for f in listFiles:
                if engine=='pandas':
                    dfNew = self.read_csv_file(containerName,directoryPath,f,engine,sourceEncoding,columnDelimiter,isFirstRowAsHeader,skipRows,skipBlankLines,addStrTechCol)
                    df = pd.concat([df, dfNew], axis=0, join="outer", ignore_index=True)
                elif engine =='polars':
                    dfNew = self.read_csv_file(containerName,directoryPath,f,engine,sourceEncoding,columnDelimiter,isFirstRowAsHeader,skipRows,skipBlankLines,addStrTechCol)
                    df = pl.concat([df, dfNew])
                    
        return df
    def read_excel_file(self,containerName:str,directoryPath:str,sourceFileName:str,engine:('pandas','polars') ='pandas',skipRows:int = 0,isFirstRowAsHeader:bool = False,sheets:list()=None,addStrTechCol:bool=False):
        if not directoryPath == '' and not directoryPath.endswith('/'):
            path = directoryPath +"/"+sourceFileName
        else:
            path = directoryPath + sourceFileName
        download_bytes = self.read_binary_file(containerName,directoryPath,sourceFileName)
        if engine=='pandas':
            if sourceFileName.endswith(".xls"):
                workbook = pd.ExcelFile(BytesIO(download_bytes), engine='xlrd') 
            else:
                workbook = pd.ExcelFile(BytesIO(download_bytes))
        elif engine=='polars':
            workbook = pl.read_excel(BytesIO(download_bytes), xlsx2csv_options={"skip_empty_lines": False},
                            read_csv_options={"has_header": isFirstRowAsHeader,"skip_rows":skipRows})
            

        workbook_sheetnames = workbook.sheet_names # get all sheet names
        if bool(sheets):
            workbook_sheetnames=list(set(workbook_sheetnames) & set(sheets))
        list_of_dff = []
        if isFirstRowAsHeader:
            isFirstRowAsHeader=0
        else:
            isFirstRowAsHeader=None
        for sheet in workbook_sheetnames:
            dff = pd.read_excel(workbook, sheet_name = sheet,skiprows=skipRows, index_col = None, header = isFirstRowAsHeader)
            if addStrTechCol:
                dff['techContainer'] = containerName
                path = path.replace("\\","/")
                dff['techFolderPath'] = path[0:path.rfind("/")]
                dff['techSourceFile'] = path[path.rfind("/")+1:len(path)]
            
            list_of_dff.append(DataFromExcel(dff,sheet))
            
        return list_of_dff
    
    def save_dataframe_as_csv(self,df:pd.DataFrame,containerName : str,directoryPath:str,file:str=None,partitionCols:list=None,sourceEncoding:str= "UTF-8", columnDelimiter:str = ";",isFirstRowAsHeader:bool = True,quoteChar:str='',quoting:[csv.QUOTE_NONE,csv.QUOTE_ALL,csv.QUOTE_MINIMAL]=csv.QUOTE_NONE,escapeChar:str="\\"):
        if not(df.empty):
            df = df.replace('\n', ' ', regex=True)
        if partitionCols:
            partitionDict = {}
            for x in partitionCols:
                partitionDict[x] = df[x].unique()
    
            partitionGroups = [dict(zip(partitionDict.keys(),items)) for items in product(*partitionDict.values())]
            partitionGroups = [d for d in partitionGroups if np.nan not in d.values()   ]

            for d in partitionGroups:
                df_part = df
                partitionPath=[]
                
                for d1 in d:
                    df_part = df_part[df_part[d1] == d[d1]]
                    partitionPath.append(f"{d1}={str(d[d1])}")
        
                if not(df_part.empty):
                    buf = BytesIO()
                    df_reset = df_part.reset_index(drop=True)
                    df_reset.to_csv(buf,index=False, sep=columnDelimiter,encoding=sourceEncoding,header=isFirstRowAsHeader,quotechar=quoteChar, quoting=quoting,escapechar=escapeChar)
                    buf.seek(0)
                    self.save_binary_file(buf.getvalue(),containerName ,directoryPath +"/" +"/".join(partitionPath),f"{uuid.uuid4().hex}.csv",True)

        else:
            if not(df.empty):
                buf = BytesIO()
                df_reset = df.reset_index(drop=True)
                df_reset.to_csv(buf,index=False, sep=columnDelimiter,encoding=sourceEncoding,header=isFirstRowAsHeader,quotechar=quoteChar, quoting=quoting,escapechar=escapeChar)
                buf.seek(0)
                if file:
                    filename = file
                else:
                    filename = f"{uuid.uuid4().hex}.csv"
                self.save_binary_file(buf.getvalue(),containerName ,directoryPath,filename,True)
   
    
    def save_dataframe_as_parquet(self,df:pd.DataFrame,containerName : str,directoryPath:str,partitionCols:list=None,compression:str=None):
        
        VALID_compression = {'snappy', 'gzip', 'brotli', None}
        if compression is not None:
            compression=compression.lower()
        if compression not in VALID_compression:
            raise ValueError("results: status must be one of %r." % VALID_compression)
        if not(df.empty):
            df = df.replace('\n', ' ', regex=True)  
        if partitionCols:
            partitionDict = {}
            for x in partitionCols:
                partitionDict[x] = df[x].unique()
    
            partitionGroups = [dict(zip(partitionDict.keys(),items)) for items in product(*partitionDict.values())]
            partitionGroups = [d for d in partitionGroups if np.nan not in d.values()   ]
            
            for d in partitionGroups:
                df_part = df
                partitionPath=[]
                
                for d1 in d:
                    df_part = df_part[df_part[d1] == d[d1]]
                    partitionPath.append(f"{d1}={str(d[d1])}")
        
                if not(df_part.empty):
                    buf = BytesIO()  
                    df_part.to_parquet(buf,allow_truncated_timestamps=True, use_deprecated_int96_timestamps=True,compression=compression)
                    buf.seek(0)
                    self.save_binary_file(buf.getvalue(),containerName ,directoryPath +"/" +"/".join(partitionPath),f"{uuid.uuid4().hex}.parquet",True)
                    
                    #new_client_parq = container_client.get_blob_client(directoryPath+"/" +"/".join(partitionPath)+"/" +f"{uuid.uuid4().hex}.parquet")
                    #new_client_parq.upload_blob(buf.getvalue(),overwrite=True)
        else:
            if not(df.empty):
                buf = BytesIO()
                df.to_parquet(buf,allow_truncated_timestamps=True, use_deprecated_int96_timestamps=True,compression=compression)
                buf.seek(0)
                self.save_binary_file(buf.getvalue(),containerName ,directoryPath,f"{uuid.uuid4().hex}.parquet",True)

                #new_client_parq = container_client.get_blob_client(directoryPath +"/"+f"{uuid.uuid4().hex}.parquet")
                #new_client_parq.upload_blob(buf.getvalue(),overwrite=True)

        
    def save_dataframe_as_parqarrow(self,df:pd.DataFrame,containerName : str,directoryPath:str,partitionCols:list=None,compression:str=None):
            
        VALID_compression = {'NONE','SNAPPY', 'GZIP', 'BROTLI', 'LZ4', 'ZSTD'}
        if compression is not None:
            compression=compression.upper()
        elif compression is None:
            compression = 'NONE'
        if not(df.empty):
            df = df.replace('\n', ' ', regex=True)
        if compression not in VALID_compression:
            raise ValueError("results: status must be one of %r." % VALID_compression)
        
        if partitionCols:
            partitionDict = {}
            for x in partitionCols:
                partitionDict[x] = df[x].unique()
    
            partitionGroups = [dict(zip(partitionDict.keys(),items)) for items in product(*partitionDict.values())]
            partitionGroups = [d for d in partitionGroups if np.nan not in d.values()   ]
            
            for d in partitionGroups:
                df_part = df
                partitionPath=[]
                
                for d1 in d:
                    df_part = df_part[df_part[d1] == d[d1]]
                    partitionPath.append(f"{d1}={str(d[d1])}")
        
                if not(df_part.empty):
                    table = pa.Table.from_pandas(df_part)
                    buf = pa.BufferOutputStream()
                    pq.write_table(table, buf,use_deprecated_int96_timestamps=True, allow_truncated_timestamps=True,compression=compression)
                    self.save_binary_file(buf.getvalue().to_pybytes(),containerName ,directoryPath +"/" +"/".join(partitionPath),f"{uuid.uuid4().hex}.parquet",True)

                    #new_client_parq = container_client.get_blob_client(directoryPath +"/" +"/".join(partitionPath)+"/"+f"{uuid.uuid4().hex}.parquet")
                    #new_client_parq.upload_blob(buf.getvalue().to_pybytes(),overwrite=True)
        else:
            if not(df.empty):
                table = pa.Table.from_pandas(df)
                buf = pa.BufferOutputStream()
                pq.write_table(table, buf,use_deprecated_int96_timestamps=True, allow_truncated_timestamps=True,compression=compression)
                self.save_binary_file(buf.getvalue().to_pybytes(),containerName ,directoryPath,f"{uuid.uuid4().hex}.parquet",True)

                #new_client_parq = container_client.get_blob_client(directoryPath +"/"+f"{uuid.uuid4().hex}.parquet")
                #new_client_parq.upload_blob(buf.getvalue().to_pybytes(),overwrite=True)

    def save_json_file(self, df: [pd.DataFrame, pl.DataFrame], containerName: str, directory: str, file:str = None, engine:['polars', 'pandas'] = 'polars', orient:['records', 'columns'] = 'records', test='x'):   

        if isinstance(df, pd.DataFrame):
            if not(df.empty):
                return
            
        if isinstance(df, pl.DataFrame):
            if not(df.is_empty):
                return    
                 
        if isinstance(df, pd.DataFrame):
            buf = BytesIO()           
            buf.seek(0)
            if engine == 'pandas':           
                df.to_json(buf, orient=orient)
            else:
                polars_df = pl.from_pandas(df)
                polars_df.write_json(buf, row_oriented=(orient=='records'), pretty=True)

        elif isinstance(df, pl.DataFrame):
            buf = BytesIO()           
            buf.seek(0)
            if engine == 'pandas':
                pandas_df = df.to_pandas()
                pandas_df.to_json(buf, orient=orient)
            else:
                df.write_json(buf, row_oriented=(orient=='records'), pretty=True)
        else:
            raise Exception("DF argument is not Pandas or Polars DataFrame")

        if file:
            filename = file
        else:
            filename = f"{uuid.uuid4().hex}.json"
        self.save_binary_file(buf.getvalue(), containerName ,directory,filename,True)

    def save_listdataframe_as_xlsx(self,list_df:list, list_sheetnames:list, containerName : str,directoryPath:str ,fileName:str,index=False,header=False):
        # for multiple dfs
        
        buf = BytesIO() 
        with pd.ExcelWriter(buf, engine = 'openpyxl') as writer:
            for df, sheet_name in zip(list_df,list_sheetnames):
                df.to_excel(writer, index = index, header = header, sheet_name = sheet_name)   
        buf.seek(0) 
        self.save_binary_file(buf.getvalue(),containerName ,directoryPath,f"{fileName}.xlsx",True)
    
    
    def read_parquet_file(self, containerName: str, directoryPath: str,sourceFileName:str, columns: list = None,addStrTechCol:bool=False)->pd.DataFrame:
        
        if not directoryPath == '' and not directoryPath.endswith('/'):
            path = directoryPath +"/"+sourceFileName
        else:
            path = directoryPath + sourceFileName
        download_bytes = self.read_binary_file(containerName,directoryPath,sourceFileName)
        df = self.read_parquet_bytes(bytes=download_bytes,columns=columns)
        
        if addStrTechCol:
            df['techContainer'] = containerName
            path = path.replace("\\","/")
            df['techFolderPath'] = path[0:path.rfind("/")]
            df['techSourceFile'] = path[path.rfind("/")+1:len(path)]
        return df
    
    def read_parquet_folder(self,containerName:str,mainDirectoryPath:str,includeSubfolders:list=None,columns:list = None,addStrTechCol:bool=False,recursive:bool=False) ->pd.DataFrame:
        
        listFiles = self.ls_files(containerName,mainDirectoryPath, recursive=recursive)
        
        if  includeSubfolders:
            newList=[]
            for i in includeSubfolders:
                for j in listFiles:
                    if  j.startswith(i.replace("/","\\")):
                        newList.append(j)
                listFiles = list(set(listFiles) - set(newList))
            listFiles = newList
            
        df = pd.DataFrame()
        if listFiles:
            for f in listFiles:
                dfNew = self.read_parquet_file(containerName,mainDirectoryPath,f,columns,addStrTechCol)
                df = pd.concat([df, dfNew], axis=0, join="outer", ignore_index=True)
        return df
    
    def delete_file(self,containerName : str,directoryPath : str,fileName:str,wait:bool=True):
        container_client = self.__service_client.get_container_client(container=containerName)
        if directoryPath =="" or directoryPath is None:
            path=fileName
        else:
            path = "/".join( (directoryPath,fileName)) 
        blob_client = container_client.get_blob_client(path)
        blob_client.delete_blob(delete_snapshots="include")
        
        if wait:
            blob_client = container_client.get_blob_client(path)
            checkIfExist = blob_client.exists()
            while checkIfExist:
                time.sleep(5)
                blob_client = container_client.get_blob_client(path)
                checkIfExist = blob_client.exists()
        
    def delete_folder(self,containerName : str,directoryPath : str,wait:bool=True):
        listFiles = self.ls_files(containerName,directoryPath,True)
        for f in listFiles:
            path = directoryPath + "/" + f.replace("\\","/")
            self.delete_file(containerName,path[0:path.rfind("/")],path[path.rfind("/")+1:len(path)])
        
        if wait:
            listFiles = self.ls_files(containerName,directoryPath,True)
            while len(listFiles)>0:
                time.sleep(5)
                listFiles = self.ls_files(containerName,directoryPath,True)
        
    def move_file(self,containerName : str,directoryPath : str,fileName:str,newContainerName : str,newDirectoryPath : str,newFileName:str,isOverWrite :bool=True,isDeleteSourceFile:bool=False):
        self.save_binary_file(self.read_binary_file(containerName,directoryPath,fileName),newContainerName,newDirectoryPath,newFileName,isOverWrite)
        if isDeleteSourceFile:
            self.delete_file(containerName ,directoryPath ,fileName)
            
    def move_folder(self,containerName : str,directoryPath : str,newContainerName : str,newDirectoryPath : str,isOverWrite :bool=True,isDeleteSourceFolder:bool=False)->bool:
        listFiles = self.ls_files(containerName,directoryPath,True)
        for i in listFiles:
            self.move_file(containerName,directoryPath,i,newContainerName,newDirectoryPath,i,True,isDeleteSourceFolder)
        return True
        
    def renema_file(self,containerName : str,directoryPath : str,fileName:str,newFileName:str):
        self.move_file(containerName,directoryPath,fileName,containerName,directoryPath,newFileName,True,True)
    
    def renema_folder(self,containerName : str,directoryPath : str,newDirectoryPath:str):
        self.move_folder(containerName,directoryPath,containerName,newDirectoryPath,True,True)
        
    def create_empty_file(self,containerName : str,directoryPath : str,fileName:str):
        container_client = self.__service_client.get_container_client(container=containerName)
        if directoryPath!='':
            directoryPath=directoryPath+'/'
        path = directoryPath +fileName
        blob_client = container_client.get_blob_client(path)
        blob_client.upload_blob('')
        
    def create_container(self,containerName : str):
        container_client = self.__service_client.get_container_client(container=containerName)   
        if not container_client.exists():
            self.__service_client.create_container(name=containerName)
        
        
    def delete_container(self,containerName : str):
        self.__service_client.delete_container(name=containerName)