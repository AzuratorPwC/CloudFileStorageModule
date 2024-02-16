#from typing import override
from multiprocessing import AuthenticationError
import csv
from shutil import ExecError
from tkinter import E
import uuid
import os
from io import BytesIO
from itertools import product
import time
import logging
import numpy as np
import pandas as pd
import polars as pl
from  openpyxl import load_workbook
from azure.storage.blob import BlobServiceClient
from azure.identity import ClientSecretCredential
from azure.identity import DefaultAzureCredential
from azure.core.exceptions import HttpResponseError,ResourceNotFoundError,ResourceExistsError
from ..Exceptions import *
import pyarrow as pa
import pyarrow.parquet as pq
from ..StorageAccountVirtualClass import StorageAccountVirtualClass
from ..Utils import *
from azure.storage.blob._shared.authentication import SharedKeyCredentialPolicy

class Blob(StorageAccountVirtualClass):
    """
    Classs
    """
    def __init__(self, url:str, access_key:str=None, tenant_id:str=None, application_id:str=None,
                 application_secret:str=None):
        super().__init__()
        try:
            if access_key is not None and tenant_id is None and application_id is None \
                and application_secret is None:
                logging.info("Blob-create by accesskey %s",url)
                self.__service_client = BlobServiceClient(account_url=url, credential=access_key,logging_enable=False)
            elif access_key is None and tenant_id is None and application_id is None \
                and application_secret is None:
                logging.info("Blob-create by defaultazurecredential %s", url)
                credential = DefaultAzureCredential()
                self.__service_client = BlobServiceClient(account_url=url, credential=credential,logging_enable=False)
            elif access_key is None and tenant_id is not None and application_id is not None \
                and application_secret is not None:
                logging.info("Blob-create by clientsecretcredential %s", url)
                token_credential = ClientSecretCredential(tenant_id, application_id,
                                                          application_secret)
                self.__service_client = BlobServiceClient(account_url=url,
                                                          credential=token_credential,logging_enable=False)
            self.__service_client.get_service_properties()
        except ResourceNotFoundError as e:
            logging.error(f"Storage account {url} not found")
            raise StorageAccountNotFound(f"Storage account {url} not found") from e
        except HttpResponseError as e:
            logging.error(f"Storage account {url} authorization error")
            raise StorageAccountAuthenticationError(f"Storage account {url} authorization error") from e
        #except 
        #except Exception as e:
        #    logging.critical(str(e))


    def check_is_dfs(self) -> bool:
        account_info = self.__service_client.get_account_information()
        return account_info['account_kind'] == 'StorageV2' and account_info['is_hns_enabled']


    def ls_files(self, container_name:str, directory_path:str, recursive:bool=False) -> list:
        try:
            container_client = self.__service_client.get_container_client(container=container_name)
            if container_client.exists() is False:
                raise ContainerNotFound(f"Container {container_name} not found")
            if not directory_path == '' and not directory_path.endswith('/'):
                directory_path += '/'

            blob_iter = container_client.list_blobs(name_starts_with=directory_path)
            files = []
            for blob in blob_iter:
                relative_path = os.path.relpath(blob.name, directory_path)
                if recursive or not '/' in relative_path:
                    files.append(relative_path)
            return files
        except ContainerNotFound as e:
            raise e
        except Exception as e:
            raise Exception(f"Error listing files in container {container_name}") from e


    def read_binary_file(self, container_name:str, directory_path:str, file_name:str) -> bytes:
        try:
            container_client = self.__service_client.get_container_client(container=container_name)
            if container_client.exists() is False:
                raise ContainerNotFound(f"Container {container_name} not found")
            if not directory_path == '' and not directory_path.endswith('/'):
                directory_path += '/'
                
            path = directory_path + file_name
            blob_client = container_client.get_blob_client(path)
            download = blob_client.download_blob()
            download_bytes = download.readall()
            return download_bytes
        except ContainerNotFound as e:
            raise e
        except ResourceNotFoundError as e:
            raise BlobNotFound(f"File {file_name} not found in container {container_name}") from e
        except Exception as e:
            raise Exception(f"Error reading file {file_name} in container {container_name}") from e
    


    def save_binary_file(self, input_bytes:bytes, container_name:str, directory_path:str,file_name:str,is_overwrite:bool=True):
        try:
            container_client = self.__service_client.get_container_client(container=container_name)
            if container_client.exists() is False:
                raise ContainerNotFound(f"Container {container_name} not found")
            if not directory_path == '' and not directory_path.endswith('/'):
                directory_path += '/'
            new_blob_client = container_client.get_blob_client(directory_path + file_name)
            new_blob_client.upload_blob(bytes(input_bytes), overwrite=is_overwrite)
        except ContainerNotFound as e:
            raise e
        except ResourceExistsError as e:
            raise BlobAlreadyExists(f"File {file_name} already exists") from e
        except Exception as e:
            raise Exception(f"Error saving file {file_name} in container {container_name}") from e


    def read_csv_file(self, container_name:str, directory_path:str, file_name:str,
                      engine:ENGINE_TYPES='polars', encoding:ENCODING_TYPES="UTF-8",
                      delimiter:DELIMITER_TYPES=',', is_first_row_as_header:bool=False,
                      skip_rows:int=0, skip_blank_lines=True,quoting:QUOTING_TYPES=None, tech_columns:bool=False):
        try:
            if not directory_path == '' and not directory_path.endswith('/'):
                path = directory_path + "/" + file_name
            else:
                path = directory_path + file_name

            download_bytes = self.read_binary_file(container_name, directory_path, file_name)
            df = self.read_csv_bytes(download_bytes, engine, encoding, delimiter,
                                    is_first_row_as_header, skip_rows, skip_blank_lines,quoting=quoting)
            if tech_columns:
                path = path.replace("\\", "/")
                df = add_tech_columns(df, container_name, path[0:path.rfind("/")],
                                  path[path.rfind("/")+1:len(path)])
            return df
        except Exception as e:
            raise e


    def read_csv_folder(self,container_name:str,directory_path:str,engine: ENGINE_TYPES = 'polars',encoding:ENCODING_TYPES = "UTF-8", delimiter:DELIMITER_TYPES = ",",is_first_row_as_header:bool = False,skip_rows:int=0,skip_blank_lines=True,quoting:QUOTING_TYPES=None,tech_columns:bool=False,recursive:bool=False):
        try:
            list_files = self.ls_files(container_name,directory_path, recursive=recursive)
            df = None
            if list_files:
                for f in list_files:
                    df_new = self.read_csv_file(container_name,directory_path,f,engine,encoding,delimiter,is_first_row_as_header,skip_rows,skip_blank_lines,quoting, tech_columns)
                    if engine=='pandas':
                        if df is None:
                            df = df_new
                        else:
                            df = pd.concat([df, df_new], axis=0, join="outer", ignore_index=True)
                    elif engine =='polars':
                        if df is None:
                            df = df_new
                        else:
                            df = pl.concat([df, df_new])
            else:
                raise FolderDataNotFound(f"Folder data {directory_path} not found in container {container_name}")
            return df
        except Exception as e:
            raise e
    def read_excel_file(self,container_name:str,directory_path:str,file_name:str,engine: ENGINE_TYPES ='polars',skip_rows:int = 0,is_first_row_as_header:bool = False,sheets:list=None,tech_columns:bool=False):
        try:
            if not directory_path == '' and not directory_path.endswith('/'):
                path = directory_path +"/"+file_name
            else:
                path = directory_path + file_name
            download_bytes = self.read_binary_file(container_name,directory_path,file_name)
            
            if file_name.endswith(".xls"):
                workbook = pd.ExcelFile(BytesIO(download_bytes), engine='xlrd')
            else:
                workbook = pd.ExcelFile(BytesIO(download_bytes))
                
            workbook_sheetnames = workbook.sheet_names # get all sheet names
            if bool(sheets):
                if  set(sheets).issubset(workbook_sheetnames) is False:
                    raise ExcelSheetNotFound(f"Sheet {sheets} not found in file {file_name}")
                    #workbook_sheetnames=list(set(workbook_sheetnames) & set(sheets))
                else:
                    load_sheets = sheets
            else:
                load_sheets = workbook_sheetnames
            
            
            list_of_dff = []
            if is_first_row_as_header:
                is_first_row_as_header=0
            else:
                is_first_row_as_header=None
            for sheet in load_sheets:
                if engine == 'pandas':
                    dff = pd.read_excel(workbook, sheet_name = sheet,skiprows=skip_rows, index_col = None, header = is_first_row_as_header)
                elif engine == 'polars':
                    dff = pl.read_excel(BytesIO(download_bytes),engine="calamine",sheet_name=sheet
                            #read_options={"has_header": is_first_row_as_header,"skip_rows":skip_rows} 
                            )
                    
                if tech_columns:
                    path = path.replace("\\","/")
                    dff =  add_tech_columns(dff,container_name,path[0:path.rfind("/")],path[path.rfind("/")+1:len(path)])
            
                list_of_dff.append(DataFromExcel(dff,sheet))
            if(len(list_of_dff)==1):
                return list_of_dff[0]
            else:
                return list_of_dff
        except Exception as e:
            raise e
    
    
    def save_dataframe_as_csv(self,df,container_name : str,directory_path:str,file_name:str=None,partition_columns:list=None,encoding:ENCODING_TYPES= "UTF-8", delimiter:DELIMITER_TYPES = ";",is_first_row_as_header:bool = True,quoting:QUOTING_TYPES=None,escape:ESCAPE_TYPES=None, engine: ENGINE_TYPES ='polars'):
        
        if isinstance(df, pd.DataFrame):
            if df.empty:
                return
            df = df.replace(r'\\n', '', regex=True)
            if engine != 'pandas':
                df = pl.from_pandas(df)

        elif isinstance(df, pl.DataFrame):
            if df.is_empty():
                return
            df = df.with_columns(pl.col(pl.Utf8).str.replace_all(r"\\n", ""))
            if engine != 'polars':
                df = df.to_pandas(use_pyarrow_extension_array=True)

                        
        if partition_columns:
            partition_dict = {}
            for x in partition_columns:
                partition_dict[x] = df[x].unique()
    
            partition_groups = [dict(zip(partition_dict.keys(),items)) for items in product(*partition_dict.values())]
            partition_groups = [d for d in partition_groups if np.nan not in d.values()]

            for d in partition_groups:
                df_part = df
                partition_path=[]

                if isinstance(df_part, pd.DataFrame):
                    for d1 in d:
                        df_part = df_part[df_part[d1] == d[d1]]
                        partition_path.append(f"{d1}={str(d[d1])}")
                    
                    if not(df_part.empty):
                        buf = BytesIO()
                        df_reset = df_part.reset_index(drop=True)
                        if quoting is not None:
                            df_reset.to_csv(buf,index=False, sep=delimiter,encoding=encoding,header=is_first_row_as_header,quotechar=quoting, quoting=1,escapechar=escape)
                        else:
                            df_reset.to_csv(buf,index=False, sep=delimiter,encoding=encoding,header=is_first_row_as_header,escapechar=escape)
                        buf.seek(0)
                        self.save_binary_file(buf.getvalue(),container_name ,directory_path +"/" + file_name + "/" +"/".join(partition_path),f"{uuid.uuid4().hex}.csv",True)


                if isinstance(df_part, pl.DataFrame):
                    for d1 in d:
                        df_part = df_part.filter(df_part[d1] == d[d1])
                        partition_path.append(f"{d1}={str(d[d1])}")
                                
                    if not(df_part.is_empty()):
                        buf = BytesIO()
                        df_reset = df_part
                        if quoting is not None:
                            df_reset.write_csv(buf, separator=delimiter, has_header=is_first_row_as_header, quote_char=quoting, quote_style='always')
                        else:
                            df_reset.write_csv(buf, separator=delimiter, has_header=is_first_row_as_header,quote_style='never')
                        buf.seek(0)
                        self.save_binary_file(buf.getvalue(),container_name ,directory_path +"/" + file_name + "/" + "/".join(partition_path),f"{uuid.uuid4().hex}.csv",True)
                                                
        else:
            buf = BytesIO()
            
            if isinstance(df, pd.DataFrame):
                df.reset_index(drop=True,inplace=True)
                if quoting is not None:
                    df.to_csv(buf,index=False, sep=delimiter,encoding=encoding,header=is_first_row_as_header,quotechar=quoting, quoting=1,escapechar=escape)
                else:
                    df.to_csv(buf,index=False, sep=delimiter,encoding=encoding,header=is_first_row_as_header,escapechar=escape)
            else:
                if quoting is not None:
                    df.write_csv(buf, separator=delimiter, has_header=is_first_row_as_header,quote_char=quoting,  quote_style="always")
                else:
                    df.write_csv(buf, separator=delimiter, has_header=is_first_row_as_header, quote_style="never")
                
    
            buf.seek(0)

            if file_name:
                file_name_check = file_name
            else:
                file_name_check  = f"{uuid.uuid4().hex}.csv"
            self.save_binary_file(buf.getvalue(),container_name ,directory_path,file_name_check,True)


    
    def save_dataframe_as_parquet(self,df,container_name : str,directory_path:str,engine: ENGINE_TYPES ='polars',partition_columns:list=None,compression:COMPRESSION_TYPES=None):
            
        if isinstance(df, pd.DataFrame):
            if df.empty:
                return
            #df = df.replace(r'\\n', '', regex=True)
            if engine != 'pandas':
                df = pl.from_pandas(df)

        elif isinstance(df, pl.DataFrame):
            if df.is_empty():
                return
            #df = df.with_columns(pl.col(pl.Utf8).str.replace_all(r"\\n", ""))
            if engine != 'polars':
                df = df.to_pandas(use_pyarrow_extension_array=True)
            
        #if  not(df.empty):
        #    df = df.replace('\n', ' ', regex=True)
        if partition_columns:
            partition_dict = {}
            for x in partition_columns:
                partition_dict[x] = df[x].unique()
    
            partition_groups = [dict(zip(partition_dict.keys(),items)) for items in product(*partition_dict.values())]
            partition_groups = [d for d in partition_groups if np.nan not in d.values()   ]
            
            for d in partition_groups:
                df_part = df
                partition_path=[]
            
            
                if isinstance(df_part, pd.DataFrame):
                    for d1 in d:
                        df_part = df_part[df_part[d1] == d[d1]]
                        partition_path.append(f"{d1}={str(d[d1])}")
                    
                    if not(df_part.empty):
                        buf = BytesIO()
                        df_part.to_parquet(buf,allow_truncated_timestamps=True, use_deprecated_int96_timestamps=True,compression=compression)
                        buf.seek(0)
                        self.save_binary_file(buf.getvalue(),container_name ,directory_path +"/" +"/".join(partition_path),f"{uuid.uuid4().hex}.parquet",True)

                if isinstance(df_part, pl.DataFrame):
                    for d1 in d:
                        df_part = df_part.filter(df_part[d1] == d[d1])
                        partition_path.append(f"{d1}={str(d[d1])}")
                                
                    if not(df_part.is_empty()):
                        buf = BytesIO()
                        df_part.write_parquet(buf,compression=compression)
                        buf.seek(0)
                        self.save_binary_file(buf.getvalue(),container_name ,directory_path +"/" + "/".join(partition_path),f"{uuid.uuid4().hex}.csv",True)
        else:
            buf = BytesIO()
            if isinstance(df, pd.DataFrame):
                df_reset = df.reset_index(drop=True)
                df_reset.to_parquet(buf,allow_truncated_timestamps=True, use_deprecated_int96_timestamps=True,compression=compression)
            else:
                df_reset = df
                df.write_parquet(buf,compression=compression)
            buf.seek(0)
            self.save_binary_file(buf.getvalue(),container_name ,directory_path,f"{uuid.uuid4().hex}.parquet",True)

        
    def save_dataframe_as_parqarrow(self,df:pd.DataFrame,container_name : str,directory_path:str,partition_columns:list=None,compression:str=None):
 
        VALID_compression = {'NONE','SNAPPY', 'GZIP', 'BROTLI', 'LZ4', 'ZSTD'}
        if compression is not None:
            compression=compression.upper()
        elif compression is None:
            compression = 'NONE'
        if not(df.empty):
            df = df.replace('\n', ' ', regex=True)
        if compression not in VALID_compression:
            raise ValueError("results: status must be one of %r." % VALID_compression)
        
        if partition_columns:
            partition_dict = {}
            for x in partition_columns:
                partition_dict[x] = df[x].unique()
    
            partition_groups = [dict(zip(partition_dict.keys(),items)) for items in product(*partition_dict.values())]
            partition_groups = [d for d in partition_groups if np.nan not in d.values()   ]
            
            for d in partition_groups:
                df_part = df
                partition_path=[]
                
                for d1 in d:
                    df_part = df_part[df_part[d1] == d[d1]]
                    partition_path.append(f"{d1}={str(d[d1])}")
        
                if not(df_part.empty):
                    table = pa.Table.from_pandas(df_part)
                    buf = pa.BufferOutputStream()
                    pq.write_table(table, buf,use_deprecated_int96_timestamps=True, allow_truncated_timestamps=True,compression=compression)
                    self.save_binary_file(buf.getvalue().to_pybytes(),container_name ,directory_path +"/" +"/".join(partition_path),f"{uuid.uuid4().hex}.parquet",True)

                    #new_client_parq = container_client.get_blob_client(directory_path +"/" +"/".join(partition_path)+"/"+f"{uuid.uuid4().hex}.parquet")
                    #new_client_parq.upload_blob(buf.getvalue().to_pybytes(),overwrite=True)
        else:
            if not(df.empty):
                table = pa.Table.from_pandas(df)
                buf = pa.BufferOutputStream()
                pq.write_table(table, buf,use_deprecated_int96_timestamps=True, allow_truncated_timestamps=True,compression=compression)
                self.save_binary_file(buf.getvalue().to_pybytes(),container_name ,directory_path,f"{uuid.uuid4().hex}.parquet",True)

                #new_client_parq = container_client.get_blob_client(directory_path +"/"+f"{uuid.uuid4().hex}.parquet")
                #new_client_parq.upload_blob(buf.getvalue().to_pybytes(),overwrite=True)

    def save_json_file(self, df, container_name: str, directory: str, file_name:str = None, engine: ENGINE_TYPES ='polars', orient:ORIENT_TYPES= 'records'):

        if isinstance(df,pd.DataFrame):
            if df.empty:
                return
            #df = df.replace(r'\\n', '', regex=True)
            if engine != 'pandas':
                df = pl.from_pandas(df)

        elif isinstance(df,pl.DataFrame):
            if df.is_empty():
                return
            #df = df.with_columns(pl.col(pl.Utf8).str.replace_all(r"\\n", ""))
            if engine != 'polars':
                df = df.to_pandas(use_pyarrow_extension_array=True)

        buf = BytesIO()    
        
        if isinstance(df, pd.DataFrame):
            df.to_json(buf, orient=orient,lines=True)
        elif isinstance(df, pl.DataFrame):
            df.write_json(buf, row_oriented=(orient=='records'), pretty=True)

        if file_name:
            file_name_check = file_name
        else:
            file_name_check = f"{uuid.uuid4().hex}.json"
        buf.seek(0)
        self.save_binary_file(buf.getvalue(), container_name ,directory,file_name_check,True)

    
    def save_dataframe_as_xlsx(self,df,container_name : str,directory_path:str ,file_name:str,sheet_name:str,engine:ENGINE_TYPES ='polars',index=False,header=False):
        if isinstance(df,pd.DataFrame):
            if df.empty:
                return
            #df = df.replace(r'\\n', '', regex=True)
            if engine != 'pandas':
                df = pl.from_pandas(df)

        elif isinstance(df,pl.DataFrame):
            if df.is_empty():
                return
            #df = df.with_columns(pl.col(pl.Utf8).str.replace_all(r"\\n", ""))
            if engine != 'polars':
                df = df.to_pandas(use_pyarrow_extension_array=True)
                
        check_if_file_exist = self.__service_client.get_container_client(container=container_name).get_blob_client(directory_path +"/" + file_name).exists()
        
        if isinstance(df,pd.DataFrame):
            if check_if_file_exist:
                excel_buf =self.read_binary_file(container_name,directory_path,file_name)
                excel_buf=BytesIO(excel_buf)
                with pd.ExcelWriter(excel_buf, engine='openpyxl',mode="a") as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index= index, header=header)
            else:
                excel_buf = BytesIO()
                with pd.ExcelWriter(excel_buf, engine = 'openpyxl') as writer:
                    df.to_excel(writer, index = index, header = header, sheet_name = sheet_name)
        elif isinstance(df,pl.DataFrame):
            if check_if_file_exist:
                excel_buf=self.read_binary_file(container_name,directory_path,file_name)
                excel_buf=BytesIO(excel_buf)
                
                df.write_excel(excel_buf, worksheet=sheet_name)
            else:
                excel_buf = BytesIO()
                df.write_excel(excel_buf, worksheet=sheet_name)
                
        excel_buf.seek(0)
        #excel_buf.close()
        self.save_binary_file(excel_buf.getvalue(),container_name ,directory_path,file_name,True)

            
    def read_parquet_file(self, container_name: str, directory_path: str,file_name:str,engine:ENGINE_TYPES ='polars', columns: list = None,tech_columns:bool=False):
        
        if not directory_path == '' and not directory_path.endswith('/'):
            path = directory_path +"/"+file_name
        else:
            path = directory_path + file_name
        download_bytes = self.read_binary_file(container_name,directory_path,file_name)
        df = self.read_parquet_bytes(input_bytes=download_bytes,columns=columns,engine=engine)
        
        if tech_columns:
            path = path.replace("\\","/")
            df =  add_tech_columns(df,container_name,path[0:path.rfind("/")],path[path.rfind("/")+1:len(path)])

        return df
    
    def read_parquet_folder(self,container_name:str,directory_path:str,engine:ENGINE_TYPES ='polars',columns:list = None,tech_columns:bool=False,recursive:bool=False):

        list_files = self.ls_files(container_name,directory_path, recursive=recursive)
            
        df = pd.DataFrame()
        if list_files:
            for f in list_files:
                df_new = self.read_parquet_file(container_name,directory_path,f,engine,columns, tech_columns)
                if engine=='pandas':
                    if df is None:
                        df = df_new
                    else:
                        df = pd.concat([df, df_new], axis=0, join="outer", ignore_index=True)
                elif engine =='polars':
                    if df is None:
                        df = df_new
                    else:
                        df = pl.concat([df, df_new])
        return df
    
    def delete_file(self,container_name : str,directory_path : str,file_name:str,wait:bool=True):
        try:
            container_client = self.__service_client.get_container_client(container=container_name)
            if container_client.exists() == False:
                raise ContainerNotFound(f"Container {container_name} not found")
            if directory_path =="" or directory_path is None:
                path=file_name
            else:
                path = "/".join( (directory_path,file_name))
            blob_client = container_client.get_blob_client(path)
            blob_client.delete_blob(delete_snapshots="include")
            if wait:
                time.sleep(1)
                blob_client = container_client.get_blob_client(path)
                check_if_exist = blob_client.exists()
                while check_if_exist:
                    time.sleep(1)
                    blob_client = container_client.get_blob_client(path)
                    check_if_exist = blob_client.exists()
        except ResourceNotFoundError as e:
            raise BlobNotFound(f"File {file_name} not found in container {container_name}") from e
        except ContainerNotFound as e:
            raise e
        except Exception as e:
            raise Exception(f"Error deleting file {file_name} in container {container_name}") from e

    def delete_files_by_prefix(self,container_name : str,directory_path : str,file_prefix:str, recursive:bool=False,wait:bool=True):
        try:
            container_client = self.__service_client.get_container_client(container=container_name)
            if container_client.exists() is False:
                raise ContainerNotFound(f"Container {container_name} not found")

            files_exists = self.ls_files(container_name,directory_path,recursive)
            files = [f for f in files_exists if f.split("/")[-1].startswith(file_prefix)]
            print(files)
            for f in files:
                blob_client = container_client.get_blob_client(directory_path+"/"+f)
                blob_client.delete_blob(delete_snapshots="include")
            if wait and files:
                time.sleep(1)
                files_exists = self.ls_files(container_name,directory_path,recursive)
                files = [f for f in files_exists if f.split("/")[-1].startswith(file_prefix)]
                while files:
                    time.sleep(1)
                    files_exists = self.ls_files(container_name,directory_path,recursive)
                    files = [f for f in files_exists if f.split("/")[-1].startswith(file_prefix)]
        except ContainerNotFound as e:
            raise e
        except Exception as e:
            raise Exception(f"Error deleting files with prefix {file_prefix} in container {container_name}") from e
        
    def delete_folder(self,container_name : str,directory_path : str,wait:bool=True):
        try:
            container_client = self.__service_client.get_container_client(container=container_name)
            if container_client.exists() is False:
                raise ContainerNotFound(f"Container {container_name} not found")
        
            list_files = self.ls_files(container_name,directory_path,True)
            for f in list_files:
                path = directory_path + "/" + f.replace("\\","/")
                self.delete_file(container_name,path[0:path.rfind("/")],path[path.rfind("/")+1:len(path)])
        
            if wait and len(list_files)>0:
                time.sleep(2)
                list_files = self.ls_files(container_name,directory_path,True)
                while len(list_files)>0:
                    time.sleep(2)
                    list_files = self.ls_files(container_name,directory_path,True)
        except ContainerNotFound as e:
            raise e
        except Exception as e:
            raise Exception(f"Error deleting folder {directory_path} in container {container_name}") from e
        
    def move_file(self,container_name : str,directory_path : str,file_name:str,new_container_name : str,new_directory_path : str,is_overwrite :bool=True,is_delete_source_file:bool=False):
        try:
            container_client = self.__service_client.get_container_client(container=container_name)
            if container_client.exists() is False:
                raise ContainerNotFound(f"Container source {container_name} not found")
            container_client = self.__service_client.get_container_client(container=new_container_name)
            if container_client.exists() is False:
                raise ContainerNotFound(f"Container target {container_name} not found")

            self.save_binary_file(self.read_binary_file(container_name,directory_path,file_name),new_container_name,new_directory_path,file_name,is_overwrite)
            if is_delete_source_file:
                self.delete_file(container_name ,directory_path ,file_name)
        except ContainerNotFound as e:
            raise e
        except ResourceNotFoundError as e:
            raise BlobNotFound(f"File {file_name} not found in container {container_name}") from e
        except Exception as e:
            raise Exception(f"Error moving file {file_name} in container {container_name} to container {new_container_name}") from e
            
    def move_folder(self,container_name : str,directory_path : str,new_container_name : str,new_directory_path : str,is_overwrite :bool=True,is_delete_source_folder:bool=False):
        try:
            container_client = self.__service_client.get_container_client(container=container_name)
            if container_client.exists() == False:
                raise ContainerNotFound(f"Container source {container_name} not found")
            container_client = self.__service_client.get_container_client(container=new_container_name)
            if container_client.exists() == False:
                raise ContainerNotFound(f"Container target {container_name} not found")
        
            list_files = self.ls_files(container_name,directory_path,True)

            for i in list_files:
                self.move_file(container_name,directory_path,i,new_container_name,new_directory_path,is_overwrite,is_delete_source_folder)
        except ContainerNotFound as e:
            raise e
        except Exception as e:
            raise Exception(f"Error moving file {file_name} in container {container_name} to container {new_container_name}") from e

        
    def renema_file(self,container_name : str,directory_path : str,file_name:str,newfile_name:str):
        try:
            container_client = self.__service_client.get_container_client(container=container_name)
            if container_client.exists() is False:
                raise ContainerNotFound(f"Container source {container_name} not found")

            self.save_binary_file(self.read_binary_file(container_name,directory_path,file_name),container_name,directory_path,newfile_name,False)
            self.delete_file(container_name ,directory_path ,file_name)
        except ContainerNotFound as e:
            raise e
        except ResourceNotFoundError as e:
            raise BlobNotFound(f"File {file_name} not found in container {container_name}") from e
        except Exception as e:
            raise Exception(f"Error moving folder {directory_path} in container {container_name}") from e
  
    def rename_folder(self,container_name : str,directory_path : str,new_directory_path:str):
        try:
            container_client = self.__service_client.get_container_client(container=container_name)
            if container_client.exists() is False:
                raise ContainerNotFound(f"Container source {container_name} not found")
        
            list_files = self.ls_files(container_name,directory_path,True)

            for i in list_files:
                self.move_file(container_name,directory_path,i,container_name,new_directory_path,False,True)
        except ContainerNotFound as e:
            raise e
        except Exception as e:
            raise Exception(f"Error moving folder {directory_path} in container {container_name}") from e
        
    def create_empty_file(self,container_name : str,directory_path:str,file_name:str):
        try:
            container_client = self.__service_client.get_container_client(container=container_name)
            if directory_path!='':
                directory_path=directory_path+'/'
            path = directory_path +file_name
            blob_client = container_client.get_blob_client(path)
            blob_client.upload_blob('')
        except ResourceNotFoundError as e:
            raise ContainerNotFound(f"Container {container_name} not found") from e
        except ResourceExistsError as e:
            raise BlobAlreadyExists(f"File {file_name} already exists") from e
        except Exception as e:
            raise Exception(f"Error creating file {file_name} in container {container_name}") from e

    def create_container(self,container_name : str,public_access:CONTAINER_ACCESS_TYPES='Private'):
        try:
            container_client = self.__service_client.get_container_client(container=container_name)
            if not container_client.exists():
                if public_access == "Private":
                    public_access = None
                self.__service_client.create_container(name=container_name,public_access=public_access)
        except ResourceExistsError as e:
            raise ContainerAccessTypes(f"Container access {public_access} is not allowe in {container_name}") from e
        except Exception as e:
            raise Exception(f"Error creating container {container_name}") from e
    def delete_container(self,container_name : str):
        try:
            container_client = self.__service_client.get_container_client(container=container_name)
            if container_client.exists():
                container_client.delete_container()
        except Exception as e:
            raise Exception(f"Error deleting container {container_name}") from e