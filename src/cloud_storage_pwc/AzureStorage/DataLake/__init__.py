
from ast import main
from ..StorageAccountVirtualClass import *

from azure.storage.filedatalake import DataLakeServiceClient

from azure.identity import DefaultAzureCredential
from azure.identity import ClientSecretCredential
import pyarrow as pa
from io import BytesIO
import pyarrow.parquet as pq
import uuid
import os
import numpy as np 
import pandas as pd
from openpyxl import Workbook
from itertools import product
import polars as pl
import time
from datetime import datetime
import csv
from ..Utils import *
from ..Exceptions import *
import logging
from azure.core.exceptions import HttpResponseError,ResourceNotFoundError,ResourceExistsError

class DataLake(StorageAccountVirtualClass):
 
    def __init__(self, url:str, access_key:str=None, tenant_id:str=None, application_id:str=None,application_secret:str=None):
        super().__init__()
        try:
            if access_key is not None and tenant_id is None and application_id is None and application_secret is None:
                logging.info("DataLake-create by accesskey %s",url)
                self.__service_client = DataLakeServiceClient(account_url=url, credential=access_key)
            elif access_key is  None and tenant_id is None and application_id is None and application_secret is None: 
                logging.info("DataLake-create by defaultazurecredential %s", url)
                credential = DefaultAzureCredential()
                self.__service_client = DataLakeServiceClient(account_url=url, credential=credential)
            elif access_key is  None and tenant_id is not None and application_id is not None and application_secret is not None:
                logging.info("DataLake-create by clientsecretcredential %s", url)
                token_credential = ClientSecretCredential(
                    tenant_id, application_id,application_secret)
                self.__service_client = DataLakeServiceClient(account_url=url, credential=token_credential)
            self.__service_client.get_service_properties()
        except ResourceNotFoundError as e:
            logging.error(f"Storage account {url} not found")
            raise StorageAccountNotFound(f"Storage account {url} not found") from e
        except HttpResponseError as e:
            logging.error(f"Storage account {url} authorization error")
            raise StorageAccountAuthenticationError(f"Storage account {url} authorization error") from e
        
    def check_is_dfs(self)->bool:
        return True

    def save_binary_file(self,input_bytes:bytes, container_name:str, directory_path:str,file_name:str,is_overwrite:bool=True):
        file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
        
        main_directory_client = file_system_client.get_directory_client("/")
        if directory_path =="":
            file_client = main_directory_client.get_file_client(file_name)
        else:
            subdir_client = main_directory_client.get_sub_directory_client(directory_path)
            if subdir_client.exists() is False:
                subdir_client.create_directory()
            file_client = subdir_client.get_file_client(file_name)
        #content_settings = ContentSettings(content_encoding=encoding,content_type = "text/csv")
        file_client.upload_data(bytes(input_bytes),overwrite=is_overwrite)
    
    def read_binary_file(self, container_name:str, directory_path:str, file_name:str)->bytes:
        file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
        
        main_directory_client = file_system_client.get_directory_client("/")
        if directory_path =="":
            file_client = main_directory_client.get_file_client(file_name)
        else:
            file_client = main_directory_client.get_sub_directory_client(directory_path).get_file_client(file_name)

        download = file_client.download_file()
        download_bytes = download.readall()
        return download_bytes

    
    def read_csv_file(self, container_name:str, directory_path:str, file_name:str,engine:ENGINE_TYPES='polars', encoding:ENCODING_TYPES="UTF-8",delimiter:DELIMITER_TYPES=',', is_first_row_as_header:bool=False,skip_rows:int=0, skip_blank_lines=True,quoting:QUOTING_TYPES=None, tech_columns:bool=False):
        
        download_bytes = self.read_binary_file(container_name,directory_path,file_name)
        df = self.read_csv_bytes(download_bytes, engine, encoding, delimiter,is_first_row_as_header, skip_rows, skip_blank_lines,quoting=quoting)
        
        if tech_columns:
            df =  add_tech_columns(df,container_name,directory_path.replace("\\","/"),file_name)
        return df
    
    def read_csv_folder(self,container_name:str,directory_path:str,engine: ENGINE_TYPES = 'polars',encoding:ENCODING_TYPES = "UTF-8", delimiter:DELIMITER_TYPES = ",",is_first_row_as_header:bool = False,skip_rows:int=0,skip_blank_lines=True,quoting:QUOTING_TYPES=None,tech_columns:bool=False,recursive:bool=False):
        
        try:
            list_files = self.ls_files(container_name,directory_path,recursive)
            #print(list_files)
            df = None
            if list_files:
                for f in list_files:
                    file = f.split("/")[-1]
                    df_new = self.read_csv_file(container_name,directory_path.removesuffix(f"/{file}"),file,engine,encoding,delimiter,is_first_row_as_header,skip_rows,skip_blank_lines,quoting,tech_columns)
                    if engine=='pandas':
                        if df is None:
                            df = df_new
                        else:
                            df = pd.concat([df, df_new], axis=0, join="outer", ignore_index=True)
                    elif engine =='polars':
                        if df is None:
                            df = df_new
                        else:
                            df = pl.concat([df, df_new])
            else:
                raise FolderDataNotFound(f"Folder data {directory_path} not found in container {container_name}")
            return df
        except Exception as e:
            raise e
        
    
    def read_excel_file(self,container_name:str,directory_path:str,file_name:str,engine: ENGINE_TYPES ='polars',skip_rows:int = 0,is_first_row_as_header:bool = False,sheets:list=None,tech_columns:bool=False):
        file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
        directory_client = file_system_client.get_directory_client(directory_path)
        file_client = directory_client.get_file_client(file_name)
        download = file_client.download_file()
        download_bytes = download.readall()
        if file_name.endswith(".xls"):
            workbook = pd.ExcelFile(download_bytes, engine='xlrd') 
        else:
            workbook = pd.ExcelFile(download_bytes)
          
        workbook_sheetnames = workbook.sheet_names # get all sheet names
        if bool(sheets):
            workbook_sheetnames=list(set(workbook_sheetnames) & set(sheets))
        list_of_dff = []
        if is_first_row_as_header:
            is_first_row_as_header=0
        else:
            is_first_row_as_header=None
        for sheet in workbook_sheetnames:
            dff = pd.read_excel(workbook, sheet_name = sheet,skip_rows=skip_rows, index_col = None, header = is_first_row_as_header)
            if tech_columns:
                df =  Utils.addTechColumns(df,container_name,directory_path.replace("\\","/"),file_name)
            
            list_of_dff.append(Utils.DataFromExcel(dff,sheet))
            
        return list_of_dff
    
    
    
    def save_dataframe_as_parquet(self,df:pd.DataFrame,container_name : str,directory_path:str,partitionCols:list=None,compression:str=None):
        
        VALID_compression = {'snappy', 'gzip', 'brotli', None}
        if compression is not None:
            compression=compression.lower()
        if compression not in VALID_compression:
            raise ValueError("results: status must be one of %r." % VALID_compression)
        if not(df.empty):
            df = df.replace('\n', ' ', regex=True)
        file_system_client = self.__service_client.get_file_system_client(file_system=container_name) 
        buf = BytesIO()
        df.to_parquet(buf,allow_truncated_timestamps=True, use_deprecated_int96_timestamps=True,partition_cols=partitionCols)
        buf.seek(0)
        directory_client = file_system_client.get_directory_client(directory_path)
        new_client_parq = directory_client.create_file(f"{uuid.uuid4().hex}.parquet")
        new_client_parq.upload_data(buf.getvalue(),overwrite=True)
    

    def save_dataframe_as_csv(self,df:[pd.DataFrame, pl.DataFrame],container_name : str,directory_path:str,file:str=None,partitionCols:list=None,encoding:ENCODING_TYPES= "UTF-8", delimiter:str = ";",is_first_row_as_header:bool = True,quoteChar:str=' ',quoting:['never', 'always', 'necessary']='never',escapeChar:str="\\", engine: ENGINE_TYPES ='polars'):
        
        quoting_dict = {'never':csv.QUOTE_NONE, 'always':csv.QUOTE_ALL, 'necessary':csv.QUOTE_MINIMAL}
        
        if isinstance(df, pd.DataFrame):
            if df.empty:
                return
            df = df.replace(r'\\n', '', regex=True) 
            if engine != 'pandas':
                df = pl.from_pandas(df)

        elif isinstance(df, pl.DataFrame):
            if df.is_empty():
                return
            df = df.with_columns(pl.col(pl.Utf8).str.replace_all(r"\\n", ""))
            if engine != 'polars':
                df = df.to_pandas(df)

                        
        if partitionCols:
            partitionDict = {}
            for x in partitionCols:
                partitionDict[x] = df[x].unique()
    
            partitionGroups = [dict(zip(partitionDict.keys(),items)) for items in product(*partitionDict.values())]
            partitionGroups = [d for d in partitionGroups if np.nan not in d.values()]

            for d in partitionGroups:
                df_part = df
                partitionPath=[]
                

                if isinstance(df_part, pd.DataFrame):
                    for d1 in d:
                        df_part = df_part[df_part[d1] == d[d1]]
                        partitionPath.append(f"{d1}={str(d[d1])}")

                if isinstance(df_part, pl.DataFrame):
                    for d1 in d:
                        df_part = df_part.filter(df_part[d1] == d[d1])
                        partitionPath.append(f"{d1}={str(d[d1])}")

                if isinstance(df_part, pd.DataFrame):
                    if not(df_part.empty):
                        buf = BytesIO()
                        df_reset = df_part.reset_index(drop=True)
                        df_reset.to_csv(buf,index=False, sep=delimiter,encoding=encoding,header=is_first_row_as_header,quotechar=quoteChar, quoting=quoting_dict[quoting],escapechar=escapeChar)
                        buf.seek(0)
                        self.save_binary_file(buf.getvalue(),container_name ,directory_path +"/" +"/".join(partitionPath),f"{uuid.uuid4().hex}.csv",True)


                if isinstance(df_part, pl.DataFrame):
                    if not(df_part.is_empty()):
                        buf = BytesIO()
                        df_reset = df_part
                        df_reset.write_csv(buf, separator=delimiter, has_header=is_first_row_as_header, quote_char='"', quote_style=quoting)
                        buf.seek(0)
                        self.save_binary_file(buf.getvalue(),container_name ,directory_path +"/" +"/".join(partitionPath),f"{uuid.uuid4().hex}.csv",True)
                                                
        
        else:
            buf = BytesIO()
            
            if isinstance(df, pd.DataFrame):
                df_reset = df.reset_index(drop=True)
                df_reset.to_csv(buf,index=False, sep=delimiter,encoding=encoding,header=is_first_row_as_header,quotechar=quoteChar, quoting=quoting_dict[quoting],escapechar=escapeChar)
            else:
                df_reset = df
                df_reset.write_csv(buf, separator=delimiter, has_header=is_first_row_as_header, quote_style=quoting)
    
            buf.seek(0)

            if file:
                filename = file
            else:
                filename = f"{uuid.uuid4().hex}.csv"
            self.save_binary_file(buf.getvalue(),container_name ,directory_path,filename,True)

        
    def save_dataframe_as_parqarrow(self,df:pd.DataFrame,container_name : str,directory_path:str,partitionCols:list=None,compression:str='NONE'):
        VALID_compression = {'NONE','SNAPPY', 'GZIP', 'BROTLI', 'LZ4', 'ZSTD'}
        if compression is not None:
            compression=compression.upper()
        elif compression is None:
            compression = 'NONE'
        if not(df.empty):
            df = df.replace('\n', ' ', regex=True)    
        if compression not in VALID_compression:
            raise ValueError("results: status must be one of %r." % VALID_compression)
        
        file_system_client = self.__service_client.get_file_system_client(file_system=container_name) 
        
        table = pa.Table.from_pandas(df)
        buf = pa.BufferOutputStream()
        pq.write_table(table, buf,use_deprecated_int96_timestamps=True, allow_truncated_timestamps=True,compression=compression)
        directory_client = file_system_client.get_directory_client(directory_path)
        new_client_parq = directory_client.create_file(f"{uuid.uuid4().hex}.parquet")
        new_client_parq.upload_data(buf.getvalue().to_pybytes(),overwrite=True)


    def save_json_file(self, df:[pd.DataFrame, pl.DataFrame], container_name: str, directory: str, file:str = None, engine: ENGINE_TYPES ='polars', orient: ORIENT_TYPES = 'records'):    

        if isinstance(df, pd.DataFrame):
            if not(df.empty):
                return
            
        if isinstance(df, pl.DataFrame):
            if not(df.is_empty):
                return    
                 
        if isinstance(df, pd.DataFrame):
            buf = BytesIO()           
            buf.seek(0)
            if engine == 'pandas':           
                df.to_json(buf, orient=orient)
            else:
                polars_df = pl.from_pandas(df)
                polars_df.write_json(buf, row_oriented=(orient=='records'), pretty=True)

        elif isinstance(df, pl.DataFrame):
            buf = BytesIO()           
            buf.seek(0)
            if engine == 'pandas':
                pandas_df = df.to_pandas()
                pandas_df.to_json(buf, orient=orient)
            else:
                df.write_json(buf, row_oriented=(orient=='records'), pretty=True)
        else:
            raise Exception("DF argument is not Pandas or Polars DataFrame")

        if file:
            filename = file
        else:
            filename = f"{uuid.uuid4().hex}.json"
        self.save_binary_file(buf.getvalue(), container_name ,directory,filename,True)

    def save_dataframe_as_xlsx(self,df,container_name : str,directory_path:str ,file_name:str,sheet_name:str,engine:ENGINE_TYPES ='polars',index=False,header=False):
        # for multiple dfs
        buf = BytesIO() 
        with pd.ExcelWriter(buf, engine = 'openpyxl') as writer:
            for df, sheet_name in zip(list_df,list_sheetnames):
                df.to_excel(writer, index = index, header = header, sheet_name = sheet_name)   
        
        file_system_client = self.__service_client.get_file_system_client(file_system=container_name) 
        directory_client = file_system_client.get_directory_client(directory_path)
        new_client_xlsx = directory_client.create_file(f"{fileName}.xlsx")
        buf.seek(0)
        new_client_xlsx.upload_data(buf.getvalue(),overwrite=True)
    
    
    def read_parquet_file(self, container_name: str, directory_path: str,file_name:str, columns: list = None,tech_columns:bool=False)->pd.DataFrame:
        file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
        directory_client = file_system_client.get_directory_client(directory_path)
        file_client= directory_client.get_file_client(file_name)
        download = file_client.download_file()
        download_bytes = download.readall()
        df = self.read_parquet_bytes(bytes=download_bytes,columns=columns)
        if tech_columns:
            df =  Utils.addTechColumns(df,container_name,directory_path.replace("\\","/"),file_name)
        
        return df
        
        
    
    def read_parquet_folder(self,container_name:str,directory_path:str,includeSubfolders:list=None,columns:list = None,tech_columns:bool=False,recursive:bool=False) ->pd.DataFrame:
        file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
        #directory_client = file_system_client.get_directory_client(directory_path)
        list_files = file_system_client.get_paths(directory_path,recursive)
        df = pd.DataFrame()
        if list_files:
            for i,r in enumerate(list_files):
                file = str(r.name).replace(directory_path + "/","")
                df_new = self.read_parquet_file(container_name,directory_path,file,columns,tech_columns)
                df = pd.concat([df, df_new], axis=0, join="outer", ignore_index=True)
        return df
    
    def delete_file(self,container_name : str,directory_path : str,fileName:str,wait:bool=True):
        file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
        
        if directory_path== "":
            file_client= file_system_client.get_file_client(fileName)
            if file_client.exists():
                file_client.delete_file()
                
                if wait:
                    checkIfExist = file_client.exists()
                    while checkIfExist:
                        time.sleep(5)
                        checkIfExist = file_client.exists()
        else:
            directory_client = file_system_client.get_directory_client(directory_path)
            if directory_client.exists():
                file_client= directory_client.get_file_client(fileName)
                if file_client.exists():
                    file_client.delete_file()
                    
                    if wait:
                        file_client= directory_client.get_file_client(fileName)
                        checkIfExist = file_client.exists()
                        while checkIfExist:
                            time.sleep(5)
                            file_client= directory_client.get_file_client(fileName)
                            checkIfExist = file_client.exists()
    def delete_files_by_prefix(self,container_name : str,directory_path : str,file_prefix:str, recursive:bool=False,wait:bool=True):
        pass
        
    def delete_folder(self,container_name : str,directory_path : str,wait:bool=True):
        file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
        if directory_path== "":
            files = file_system_client.get_paths()
            for file in files:
                file.delete_file()
                
            if wait:
                files = list(file_system_client.get_paths())
                while len(files)>0:
                    files = file_system_client.get_paths()      
        else:
                 
            directory_client = file_system_client.get_directory_client(directory_path)
            if directory_client.exists():
                directory_client.delete_directory()
        
            if wait:
                directory_client = file_system_client.get_directory_client(directory_path)
                checkIfExist=directory_client.exists()
                while checkIfExist:
                    time.sleep(5)
                    directory_client = file_system_client.get_directory_client(directory_path)
                    checkIfExist=directory_client.exists()
    
    def file_exists(self, container_name : str,directory_path : str,fileName:str):
        file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
        
        if directory_path== "":
            file_client= file_system_client.get_file_client(fileName)
        else:
            directory_client = file_system_client.get_directory_client(directory_path)
            if directory_client.exists():
                file_client= directory_client.get_file_client(fileName)
        return file_client.exists()
            
    def folder_exists(self, container_name : str, directory_path : str):
        if directory_path== "":
            return False       
        file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
        directory_client = file_system_client.get_directory_client(directory_path)
        return directory_client.exists()
                    
    
    def move_file(self,container_name : str,directory_path : str,fileName:str,newcontainer_name : str,newdirectory_path : str,newFileName:str,isOverWrite :bool=True,isDeleteSourceFile:bool=False):
        if not self.file_exists(container_name, directory_path, fileName):
            raise FileNotFoundError
        self.save_binary_file(self.read_binary_file(container_name,directory_path,fileName),newcontainer_name,newdirectory_path,newFileName,isOverWrite)
        if isDeleteSourceFile:
            self.delete_file(container_name ,directory_path ,fileName)
    
    def ls_files(self, container_name:str, directory_path:str, recursive:bool=False)->list:
        try:
            file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
            if file_system_client.exists() is False:
                raise ContainerNotFound(f"Container {container_name} not found")

            if directory_path=="":
                directory_path="/"
            files = []
            generator = file_system_client.get_paths(path=directory_path, recursive=recursive)
            for file in generator:
                if file.is_directory is False:
                    files.append(file.name)
            return files
        except ContainerNotFound as e:
            raise e
        except Exception as e:
            raise Exception(f"Error listing files in container {container_name}") from e
        
        
    
    def move_folder(self,container_name : str,directory_path : str,newcontainer_name : str,newdirectory_path : str,isOverWrite :bool=True,isDeleteSourceFolder:bool=False)->bool:
        if not self.folder_exists(container_name, directory_path):
            raise FileNotFoundError
        list_files = self.ls_files(container_name,directory_path,True)
        for i in list_files:
            self.move_file(container_name,directory_path,i,newcontainer_name,newdirectory_path,i,isOverWrite,isDeleteSourceFolder)
        return True
        
    def renema_file(self,container_name : str,directory_path : str,fileName:str,newFileName:str):
        self.move_file(container_name,directory_path,fileName,container_name,directory_path,newFileName,False,True)
    
    def rename_folder(self,container_name : str,directory_path : str,new_directory_path:str):
        pass
        #self.move_folder(container_name,directory_path,container_name,newdirectory_path,False,True)
        
    def create_empty_file(self,container_name : str,directory_path : str,fileName:str):
        file_system_client = self.__service_client.get_file_system_client(file_system=container_name)
        directory_client = file_system_client.get_directory_client(directory_path)
        file_client= directory_client.get_file_client(fileName)
        file_client.create_file()
        file_client.upload_data('')
        
    def create_container(self,container_name : str,public_access:CONTAINER_ACCESS_TYPES='Private'):
        try:
            container_client = self.__service_client.get_file_system_client(file_system=container_name)
            if not container_client.exists():
                if public_access == "Private":
                    public_access = None
                self.__service_client.create_file_system(file_system=container_name,public_access=public_access)
        except ResourceExistsError as e:
            raise ContainerAccessTypes(f"Container access {public_access} is not allowe in {container_name}") from e
        except Exception as e:
            raise Exception(f"Error creating container {container_name}") from e

        
    def delete_container(self,container_name : str):
        try:
            container_client = self.__service_client.get_file_system_client(file_system=container_name)
            if container_client.exists():
                container_client.delete_file_system()
        except Exception as e:
            raise Exception(f"Error deleting container {container_name}") from e